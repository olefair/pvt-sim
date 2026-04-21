"""Consolidated GUI widget contract tests for PVT-SIM desktop application.

Replaces:
  test_pvtapp_desktop_contract.py
  test_pvtapp_conditions_input.py
  test_pvtapp_zero_fraction_duplicates.py
  test_pvtapp_cvd_result_views.py
  test_pvtapp_tbp_result_views.py
  test_pvtapp_pt_flash_viscosity.py
  test_pvtapp_phase_envelope_widget_style.py
  test_pvtapp_stability_desktop.py
  test_pvtapp_workspace_layout.py
"""

from __future__ import annotations

import csv
import gc
import json
import os
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import pytest

pytestmark = pytest.mark.gui_contract

try:
    from PySide6.QtCore import QCoreApplication, QEvent, QSettings, Qt
    from PySide6.QtGui import QColor
    from PySide6.QtTest import QTest
    from PySide6.QtWidgets import (
        QAbstractItemView,
        QApplication,
        QLabel,
        QLineEdit,
        QMessageBox,
        QSplitter,
        QStyle,
        QStyleOptionComboBox,
        QStyleOptionViewItem,
        QWidget,
    )
except ModuleNotFoundError:  # pragma: no cover
    QCoreApplication = None  # type: ignore[assignment]
    QEvent = None  # type: ignore[assignment]
    QSettings = None  # type: ignore[assignment]
    Qt = None  # type: ignore[assignment]
    QColor = None  # type: ignore[assignment]
    QTest = None  # type: ignore[assignment]
    QAbstractItemView = None  # type: ignore[assignment]
    QApplication = None  # type: ignore[assignment]
    QLabel = None  # type: ignore[assignment]
    QLineEdit = None  # type: ignore[assignment]
    QMessageBox = None  # type: ignore[assignment]
    QSplitter = None  # type: ignore[assignment]
    QStyle = None  # type: ignore[assignment]
    QStyleOptionComboBox = None  # type: ignore[assignment]
    QStyleOptionViewItem = None  # type: ignore[assignment]
    QWidget = None  # type: ignore[assignment]

from pvtapp.capabilities import (
    GUI_CALCULATION_TYPE_LABELS,
    GUI_EOS_TYPE_LABELS,
    GUI_SUPPORTED_CALCULATION_TYPES,
    GUI_SUPPORTED_EOS_TYPES,
)
from pvtapp.schemas import (
    BubblePointResult,
    CalculationType,
    CCEResult,
    CCEStepResult,
    ConvergenceStatusEnum,
    CVDResult,
    CVDStepResult,
    DewPointResult,
    DLResult,
    DLStepResult,
    EOSType,
    FluidComposition,
    IterationRecord,
    PhaseEnvelopePoint,
    PhaseEnvelopeResult,
    PhaseEnvelopeTracingMethod,
    PlusFractionCharacterizationPreset,
    PressureUnit,
    PTFlashResult,
    RunConfig,
    RunResult,
    RunStatus,
    SeparatorResult,
    SeparatorStageResult,
    SolverDiagnostics,
    SolverSettings,
    StabilityAnalysisResult,
    StabilitySeedResultData,
    StabilityTrialResultData,
    TBPCharacterizationContext,
    TBPCharacterizationCutMapping,
    TBPCharacterizationPedersenFit,
    TBPCharacterizationSCNEntry,
    TBPExperimentCutResult,
    TBPExperimentResult,
    TemperatureUnit,
    pressure_from_pa,
    temperature_from_k,
)
from pvtapp.style import DEFAULT_UI_SCALE, UI_SCALE_STEP, build_cato_stylesheet, scale_metric
from pvtcore.models import resolve_component_id
from pvtcore.validation.pete665_assignment import psia_to_pa

try:
    from pvtapp.assignment_case import build_assignment_desktop_preset
    from pvtapp.component_catalog import STANDARD_COMPONENTS
    from pvtapp.job_runner import run_calculation
    from pvtapp.main import PVTSimulatorWindow
    from pvtapp.widgets.composition_input import (
        COMPONENT_DROPDOWN_BUTTON_WIDTH,
        COMPONENT_PICKER_OPTIONS,
        HEAVY_MODE_INLINE,
        HEAVY_MODE_PLUS,
        MOLE_FRACTION_COLUMN_MIN_WIDTH,
        CompositionInputWidget,
    )
    from pvtapp.widgets.conditions_input import ConditionsInputWidget
    from pvtapp.widgets.diagnostics_view import DiagnosticsWidget
    from pvtapp.widgets.results_view import (
        PLOT_CANVAS_COLOR,
        PLOT_SURFACE_COLOR,
        ResultsPlotWidget,
        ResultsSidebarWidget,
        ResultsTableWidget,
    )
    from pvtapp.widgets.run_log_view import RunLogWidget
    from pvtapp.widgets.text_output_view import TextOutputWidget
    from pvtapp.widgets.two_pane_workspace import TwoPaneWorkspace, ViewSpec
except ModuleNotFoundError:  # pragma: no cover
    build_assignment_desktop_preset = None  # type: ignore[assignment]
    STANDARD_COMPONENTS = None  # type: ignore[assignment]
    run_calculation = None  # type: ignore[assignment]
    PVTSimulatorWindow = None  # type: ignore[assignment]
    COMPONENT_DROPDOWN_BUTTON_WIDTH = None  # type: ignore[assignment]
    COMPONENT_PICKER_OPTIONS = None  # type: ignore[assignment]
    HEAVY_MODE_INLINE = None  # type: ignore[assignment]
    HEAVY_MODE_PLUS = None  # type: ignore[assignment]
    MOLE_FRACTION_COLUMN_MIN_WIDTH = None  # type: ignore[assignment]
    CompositionInputWidget = None  # type: ignore[assignment]
    ConditionsInputWidget = None  # type: ignore[assignment]
    DiagnosticsWidget = None  # type: ignore[assignment]
    PLOT_CANVAS_COLOR = None  # type: ignore[assignment]
    PLOT_SURFACE_COLOR = None  # type: ignore[assignment]
    ResultsPlotWidget = None  # type: ignore[assignment]
    ResultsSidebarWidget = None  # type: ignore[assignment]
    ResultsTableWidget = None  # type: ignore[assignment]
    RunLogWidget = None  # type: ignore[assignment]
    TextOutputWidget = None  # type: ignore[assignment]
    TwoPaneWorkspace = None  # type: ignore[assignment]
    ViewSpec = None  # type: ignore[assignment]


# ── Fixtures ──────────────────────────────────────────────────────────────


@pytest.fixture(scope="module")
def app() -> QApplication:
    if (
        QApplication is None
        or PVTSimulatorWindow is None
        or DiagnosticsWidget is None
        or ResultsPlotWidget is None
        or ResultsTableWidget is None
        or TextOutputWidget is None
    ):
        pytest.skip("PySide6/matplotlib is not installed in this test environment")
    instance = QApplication.instance()
    if instance is not None:
        return instance
    return QApplication([])


@pytest.fixture()
def settings_path(tmp_path: Path) -> Path:
    return tmp_path / "pvtapp-test-settings.ini"


@pytest.fixture()
def window(app: QApplication, monkeypatch: pytest.MonkeyPatch, settings_path: Path) -> PVTSimulatorWindow:
    def _create_settings(_self) -> QSettings:
        return QSettings(str(settings_path), QSettings.Format.IniFormat)

    monkeypatch.setattr(PVTSimulatorWindow, "_create_settings", _create_settings)
    instance = PVTSimulatorWindow()
    yield instance
    instance.close()
    instance.deleteLater()
    app.processEvents()
    if QCoreApplication is not None and QEvent is not None:
        QCoreApplication.sendPostedEvents(None, QEvent.Type.DeferredDelete.value)
        app.processEvents()
    gc.collect()


# ── RunResult factory helpers ─────────────────────────────────────────────


def _run_config(data: dict) -> RunConfig:
    return RunConfig.model_validate(data)


def _ts() -> datetime:
    return datetime(2026, 4, 11, 9, 0, 0)


def _completed(config: RunConfig, **payload) -> RunResult:
    return RunResult(
        run_id=f"{config.calculation_type.value}-result",
        run_name=f"{config.calculation_type.value}-result",
        status=RunStatus.COMPLETED,
        started_at=_ts(),
        completed_at=_ts(),
        duration_seconds=1.0,
        config=config,
        **payload,
    )


def _simple_diag(iters: int = 4, residual: float = 1e-12) -> SolverDiagnostics:
    return SolverDiagnostics(
        status=ConvergenceStatusEnum.CONVERGED,
        iterations=iters,
        final_residual=residual,
    )


# ── Per-calc-type config/result factories ─────────────────────────────────

def _pt_flash_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.55},
            {"component_id": "C10", "mole_fraction": 0.45},
        ]},
        "calculation_type": "pt_flash",
        "eos_type": "peng_robinson",
        "pt_flash_config": {"pressure_pa": 8.0e6, "temperature_k": 350.0},
    })


def _pt_flash_res() -> RunResult:
    return _completed(
        _pt_flash_cfg(),
        pt_flash_result=PTFlashResult(
            converged=True, phase="two-phase", vapor_fraction=0.35,
            liquid_composition={"C1": 0.25, "C10": 0.75},
            vapor_composition={"C1": 0.92, "C10": 0.08},
            K_values={"C1": 3.68, "C10": 0.11},
            liquid_fugacity={"C1": 1.0, "C10": 1.0},
            vapor_fugacity={"C1": 1.0, "C10": 1.0},
            diagnostics=_simple_diag(),
        ),
    )


def _bubble_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.50},
            {"component_id": "C10", "mole_fraction": 0.50},
        ]},
        "calculation_type": "bubble_point",
        "eos_type": "peng_robinson",
        "bubble_point_config": {"temperature_k": 350.0, "pressure_initial_pa": 1.25e7},
    })


def _bubble_res() -> RunResult:
    return _completed(
        _bubble_cfg(),
        bubble_point_result=BubblePointResult(
            converged=True, pressure_pa=1.20e7, temperature_k=350.0,
            iterations=5, residual=1.0e-10, stable_liquid=True,
            liquid_composition={"C1": 0.50, "C10": 0.50},
            vapor_composition={"C1": 0.88, "C10": 0.12},
            k_values={"C1": 1.76, "C10": 0.24},
            diagnostics=_simple_diag(5, 1e-10),
        ),
    )


def _bubble_plus_cfg() -> RunConfig:
    return _run_config({
        "composition": {
            "components": [
                {"component_id": "N2", "mole_fraction": 0.0021},
                {"component_id": "CO2", "mole_fraction": 0.0187},
                {"component_id": "C1", "mole_fraction": 0.3478},
                {"component_id": "C2", "mole_fraction": 0.0712},
                {"component_id": "C3", "mole_fraction": 0.0934},
                {"component_id": "iC4", "mole_fraction": 0.0302},
                {"component_id": "nC4", "mole_fraction": 0.0431},
                {"component_id": "iC5", "mole_fraction": 0.0276},
                {"component_id": "nC5", "mole_fraction": 0.0418},
                {"component_id": "C6", "mole_fraction": 0.0574},
            ],
            "plus_fraction": {
                "label": "C7+", "cut_start": 7, "z_plus": 0.2667,
                "mw_plus_g_per_mol": 119.787599, "sg_plus_60f": 0.82,
                "characterization_preset": "manual", "max_carbon_number": 20,
                "split_method": "katz", "split_mw_model": "table",
                "lumping_enabled": True, "lumping_n_groups": 6,
            },
        },
        "calculation_type": "bubble_point",
        "eos_type": "peng_robinson",
        "bubble_point_config": {
            "temperature_k": 360.0, "pressure_initial_pa": 1.0e5,
            "pressure_unit": "bar", "temperature_unit": "C",
        },
    })


def _pt_flash_plus_tbp_fit_cfg() -> RunConfig:
    return _run_config({
        "composition": {
            "components": [
                {"component_id": "C1", "mole_fraction": 0.35},
                {"component_id": "C2", "mole_fraction": 0.20},
                {"component_id": "C3", "mole_fraction": 0.15},
            ],
            "plus_fraction": {
                "label": "C7+", "cut_start": 7, "z_plus": 0.30,
                "mw_plus_g_per_mol": 108.13333333333333, "sg_plus_60f": 0.82,
                "characterization_preset": "manual", "max_carbon_number": 12,
                "split_method": "pedersen", "split_mw_model": "paraffin",
                "pedersen_solve_ab_from": "fit_to_tbp",
                "tbp_cuts": [
                    {"name": "C7", "z": 0.120, "mw": 96.0},
                    {"name": "C8", "z": 0.100, "mw": 110.0},
                    {"name": "C9", "z": 0.080, "mw": 124.0, "tb_k": 425.0},
                ],
            },
        },
        "calculation_type": "pt_flash",
        "eos_type": "peng_robinson",
        "pt_flash_config": {"pressure_pa": 5.0e6, "temperature_k": 350.0},
    })


def _inline_pseudo_bubble_cfg() -> RunConfig:
    return _run_config({
        "composition": {
            "components": [
                {"component_id": "C1", "mole_fraction": 0.199620},
                {"component_id": "C2", "mole_fraction": 0.100100},
                {"component_id": "C3", "mole_fraction": 0.185790},
                {"component_id": "nC4", "mole_fraction": 0.090360},
                {"component_id": "nC5", "mole_fraction": 0.188510},
                {"component_id": "PSEUDO_PLUS", "mole_fraction": 0.235630},
            ],
            "inline_components": [{
                "component_id": "PSEUDO_PLUS", "name": "Pseudo+",
                "formula": "Pseudo+", "molecular_weight_g_per_mol": 86.177,
                "critical_temperature_k": 507.4, "critical_pressure_pa": 3008134.215801,
                "omega": 0.296,
            }],
        },
        "calculation_type": "bubble_point",
        "eos_type": "peng_robinson",
        "bubble_point_config": {
            "temperature_k": 326.76111111111106, "pressure_initial_pa": 1.0e7,
            "pressure_unit": "bar", "temperature_unit": "F",
        },
    })


def _inline_pseudo_bubble_res() -> RunResult:
    return _completed(
        _inline_pseudo_bubble_cfg(),
        bubble_point_result=BubblePointResult(
            converged=True, pressure_pa=5.236495885582632e6,
            temperature_k=326.76111111111106, iterations=12,
            residual=6.16e-16, stable_liquid=True,
            liquid_composition={
                "C1": 0.199618, "C2": 0.100099, "C3": 0.185788,
                "nC4": 0.090359, "nC5": 0.188508, "PSEUDO_PLUS": 0.235628,
            },
            vapor_composition={
                "C1": 0.724185, "C2": 0.120726, "C3": 0.099516,
                "nC4": 0.021807, "nC5": 0.021056, "PSEUDO_PLUS": 0.012711,
            },
            k_values={
                "C1": 3.627854, "C2": 1.206062, "C3": 0.535642,
                "nC4": 0.241339, "nC5": 0.111696, "PSEUDO_PLUS": 0.053944,
            },
            diagnostics=_simple_diag(12, 6.16e-16),
        ),
    )


def _dew_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.85},
            {"component_id": "C3", "mole_fraction": 0.10},
            {"component_id": "C7", "mole_fraction": 0.05},
        ]},
        "calculation_type": "dew_point",
        "eos_type": "peng_robinson",
        "dew_point_config": {"temperature_k": 380.0, "pressure_initial_pa": 2.10e7},
    })


def _dew_res() -> RunResult:
    return _completed(
        _dew_cfg(),
        dew_point_result=DewPointResult(
            converged=True, pressure_pa=1.95e7, temperature_k=380.0,
            iterations=6, residual=8.0e-11, stable_vapor=True,
            liquid_composition={"C1": 0.52, "C3": 0.28, "C7": 0.20},
            vapor_composition={"C1": 0.85, "C3": 0.10, "C7": 0.05},
            k_values={"C1": 1.63, "C3": 0.82, "C7": 0.21},
            diagnostics=_simple_diag(6, 8e-11),
        ),
    )


def _envelope_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.60},
            {"component_id": "C4", "mole_fraction": 0.25},
            {"component_id": "C10", "mole_fraction": 0.15},
        ]},
        "calculation_type": "phase_envelope",
        "eos_type": "peng_robinson",
        "phase_envelope_config": {
            "temperature_min_k": 250.0, "temperature_max_k": 420.0, "n_points": 24,
        },
    })


def _envelope_res() -> RunResult:
    return _completed(
        _envelope_cfg(),
        phase_envelope_result=PhaseEnvelopeResult(
            bubble_curve=[
                PhaseEnvelopePoint(temperature_k=300.0, pressure_pa=8.0e6, point_type="bubble"),
                PhaseEnvelopePoint(temperature_k=320.0, pressure_pa=7.0e6, point_type="bubble"),
            ],
            dew_curve=[
                PhaseEnvelopePoint(temperature_k=340.0, pressure_pa=6.5e6, point_type="dew"),
                PhaseEnvelopePoint(temperature_k=360.0, pressure_pa=5.5e6, point_type="dew"),
            ],
            critical_point=PhaseEnvelopePoint(temperature_k=330.0, pressure_pa=7.2e6, point_type="critical"),
        ),
    )


def _tbp_cfg() -> RunConfig:
    return _run_config({
        "calculation_type": "tbp",
        "tbp_config": {
            "cut_start": 7,
            "cuts": [
                {"name": "C7", "z": 0.020, "mw": 96.0, "sg": 0.74},
                {"name": "C8", "z": 0.015, "mw": 110.0, "sg": 0.77},
                {"name": "C9", "z": 0.015, "mw": 124.0, "sg": 0.80},
            ],
        },
    })


def _tbp_res() -> RunResult:
    return RunResult(
        run_id="tbp-view-test",
        run_name="tbp-view-test",
        status=RunStatus.COMPLETED,
        started_at=datetime(2026, 4, 14, 11, 0, 0),
        completed_at=datetime(2026, 4, 14, 11, 0, 2),
        duration_seconds=2.0,
        config=RunConfig(
            calculation_type=CalculationType.TBP,
            solver_settings=SolverSettings(),
            tbp_config={
                "cuts": [
                    {"name": "C7-C9", "z": 0.020, "mw": 103.0, "sg": 0.74, "tb_k": 385.0},
                    {"name": "C12", "z": 0.015, "mw": 170.0, "sg": 0.77},
                    {"name": "C15-C18", "z": 0.015, "mw": 235.0, "sg": 0.80},
                ]
            },
        ),
        tbp_result=TBPExperimentResult(
            cut_start=7, cut_end=18, z_plus=0.05, mw_plus_g_per_mol=164.8,
            cuts=[
                TBPExperimentCutResult(
                    name="C7-C9", carbon_number=7, carbon_number_end=9,
                    mole_fraction=0.020, normalized_mole_fraction=0.4,
                    cumulative_mole_fraction=0.4, molecular_weight_g_per_mol=103.0,
                    normalized_mass_fraction=0.25, cumulative_mass_fraction=0.25,
                    specific_gravity=0.74, boiling_point_k=385.0,
                    boiling_point_source="input",
                ),
                TBPExperimentCutResult(
                    name="C12", carbon_number=12, carbon_number_end=12,
                    mole_fraction=0.015, normalized_mole_fraction=0.3,
                    cumulative_mole_fraction=0.7, molecular_weight_g_per_mol=170.0,
                    normalized_mass_fraction=0.3098305084745763,
                    cumulative_mass_fraction=0.5598305084745763,
                    specific_gravity=0.77, boiling_point_k=500.0,
                    boiling_point_source="estimated_soreide",
                ),
                TBPExperimentCutResult(
                    name="C15-C18", carbon_number=15, carbon_number_end=18,
                    mole_fraction=0.015, normalized_mole_fraction=0.3,
                    cumulative_mole_fraction=1.0, molecular_weight_g_per_mol=235.0,
                    normalized_mass_fraction=0.4401694915254237,
                    cumulative_mass_fraction=1.0,
                    specific_gravity=0.80, boiling_point_k=620.0,
                    boiling_point_source="estimated_soreide",
                ),
            ],
            characterization_context=TBPCharacterizationContext(
                bridge_status="characterized_scn",
                plus_fraction_label="C7+", cut_start=7, cut_end=18,
                cut_count=3, z_plus=0.05, mw_plus_g_per_mol=164.8,
                characterization_method="pedersen_fit_to_tbp",
                split_mw_model="table",
                pseudo_property_correlation="riazi_daubert",
                runtime_component_basis="scn_unlumped",
                pedersen_fit=TBPCharacterizationPedersenFit(
                    solve_ab_from="fit_to_tbp", A=-3.2571, B=0.0124,
                    tbp_cut_rms_relative_error=0.0182,
                ),
                cut_mappings=[
                    TBPCharacterizationCutMapping(
                        cut_name="C7-C9", carbon_number=7, carbon_number_end=9,
                        observed_mole_fraction=0.020, observed_normalized_mole_fraction=0.4,
                        characterized_mole_fraction=0.0202,
                        characterized_normalized_mole_fraction=0.404,
                        characterized_average_molecular_weight_g_per_mol=104.0,
                        normalized_relative_error=0.01, scn_members=[7, 8, 9],
                    ),
                    TBPCharacterizationCutMapping(
                        cut_name="C12", carbon_number=12, carbon_number_end=12,
                        observed_mole_fraction=0.015, observed_normalized_mole_fraction=0.3,
                        characterized_mole_fraction=0.0148,
                        characterized_normalized_mole_fraction=0.296,
                        characterized_average_molecular_weight_g_per_mol=170.0,
                        normalized_relative_error=-0.0133, scn_members=[12],
                    ),
                ],
                scn_distribution=[
                    TBPCharacterizationSCNEntry(
                        component_id="SCN7", carbon_number=7,
                        assay_mole_fraction=0.0067, normalized_mole_fraction=0.134,
                        normalized_mass_fraction=0.079, molecular_weight_g_per_mol=96.0,
                        specific_gravity_60f=0.731, boiling_point_k=371.6,
                        critical_temperature_k=540.2, critical_pressure_pa=2740000.0,
                        critical_volume_m3_per_mol=0.00043, omega=0.351,
                    ),
                    TBPCharacterizationSCNEntry(
                        component_id="SCN12", carbon_number=12,
                        assay_mole_fraction=0.0148, normalized_mole_fraction=0.296,
                        normalized_mass_fraction=0.305, molecular_weight_g_per_mol=170.0,
                        specific_gravity_60f=0.781, boiling_point_k=489.2,
                        critical_temperature_k=665.4, critical_pressure_pa=1820000.0,
                        critical_volume_m3_per_mol=0.00071, omega=0.615,
                    ),
                    TBPCharacterizationSCNEntry(
                        component_id="SCN15", carbon_number=15,
                        assay_mole_fraction=0.0096, normalized_mole_fraction=0.192,
                        normalized_mass_fraction=0.287, molecular_weight_g_per_mol=212.0,
                        specific_gravity_60f=0.812, boiling_point_k=572.8,
                        critical_temperature_k=742.1, critical_pressure_pa=1460000.0,
                        critical_volume_m3_per_mol=0.00092, omega=0.792,
                    ),
                ],
                notes=[
                    "Standalone TBP assay artifacts now preserve a bounded SCN characterization bridge.",
                ],
            ),
        ),
    )


def _cce_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.70},
            {"component_id": "C4", "mole_fraction": 0.20},
            {"component_id": "C10", "mole_fraction": 0.10},
        ]},
        "calculation_type": "cce",
        "eos_type": "peng_robinson",
        "cce_config": {
            "temperature_k": 360.0, "pressure_start_pa": 2.0e7,
            "pressure_end_pa": 2.0e6, "n_steps": 6,
        },
    })


def _cce_res() -> RunResult:
    return _completed(
        _cce_cfg(),
        cce_result=CCEResult(
            temperature_k=360.0, saturation_pressure_pa=1.55e7,
            steps=[
                CCEStepResult(
                    pressure_pa=2.0e7, relative_volume=1.00,
                    liquid_fraction=1.0, vapor_fraction=0.0, z_factor=0.82,
                    liquid_density_kg_per_m3=648.2, vapor_density_kg_per_m3=None,
                    liquid_viscosity_pa_s=0.0018, vapor_viscosity_pa_s=None,
                ),
                CCEStepResult(
                    pressure_pa=1.2e7, relative_volume=1.15,
                    liquid_fraction=0.72, vapor_fraction=0.28, z_factor=0.91,
                    liquid_density_kg_per_m3=583.4, vapor_density_kg_per_m3=128.6,
                    liquid_viscosity_pa_s=0.0009, vapor_viscosity_pa_s=0.00002,
                ),
            ],
        ),
    )


def _dl_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.40},
            {"component_id": "C3", "mole_fraction": 0.30},
            {"component_id": "C10", "mole_fraction": 0.30},
        ]},
        "calculation_type": "differential_liberation",
        "eos_type": "peng_robinson",
        "dl_config": {
            "temperature_k": 350.0, "bubble_pressure_pa": 1.5e7,
            "pressure_end_pa": 1.0e6, "n_steps": 8,
        },
    })


def _dl_res() -> RunResult:
    return _completed(
        _dl_cfg(),
        dl_result=DLResult(
            temperature_k=350.0, bubble_pressure_pa=1.5e7,
            rsi=620.0, boi=1.48, residual_oil_density_kg_per_m3=782.0,
            converged=True,
            steps=[
                DLStepResult(
                    pressure_pa=1.5e7, rs=620.0, bg=None, bo=1.48, bt=1.48,
                    vapor_fraction=0.0, oil_density_kg_per_m3=648.2,
                    oil_viscosity_pa_s=0.0015, gas_gravity=None, gas_z_factor=None,
                    gas_viscosity_pa_s=None, cumulative_gas_produced=0.0,
                    liquid_moles_remaining=1.0,
                ),
                DLStepResult(
                    pressure_pa=5.0e6, rs=210.0, bg=0.0048, bo=1.18, bt=1.23,
                    vapor_fraction=0.28, oil_density_kg_per_m3=701.4,
                    oil_viscosity_pa_s=0.0011, gas_gravity=0.8123, gas_z_factor=0.9456,
                    gas_viscosity_pa_s=0.000018, cumulative_gas_produced=410.0,
                    liquid_moles_remaining=0.72,
                ),
            ],
        ),
    )


def _cvd_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.85},
            {"component_id": "C3", "mole_fraction": 0.10},
            {"component_id": "C7", "mole_fraction": 0.05},
        ]},
        "calculation_type": "cvd",
        "eos_type": "peng_robinson",
        "cvd_config": {
            "temperature_k": 380.0, "dew_pressure_pa": 5.652e6,
            "pressure_end_pa": 5.0e6, "n_steps": 5,
        },
    })


def _cvd_res() -> RunResult:
    return _completed(
        _cvd_cfg(),
        cvd_result=CVDResult(
            temperature_k=380.0, dew_pressure_pa=5.652e6,
            initial_z=0.92, converged=True,
            steps=[
                CVDStepResult(
                    pressure_pa=5.652e6, liquid_dropout=0.00, gas_produced=0.00,
                    cumulative_gas_produced=0.00, moles_remaining=1.0,
                    z_two_phase=0.92, liquid_density_kg_per_m3=None,
                    vapor_density_kg_per_m3=120.0, liquid_viscosity_pa_s=None,
                    vapor_viscosity_pa_s=0.000015,
                ),
                CVDStepResult(
                    pressure_pa=5.0e6, liquid_dropout=0.07, gas_produced=0.20,
                    cumulative_gas_produced=0.20, moles_remaining=0.80,
                    z_two_phase=0.86, liquid_density_kg_per_m3=530.0,
                    vapor_density_kg_per_m3=96.0, liquid_viscosity_pa_s=0.00055,
                    vapor_viscosity_pa_s=0.000013,
                ),
            ],
        ),
    )


def _separator_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.40},
            {"component_id": "C4", "mole_fraction": 0.35},
            {"component_id": "C10", "mole_fraction": 0.25},
        ]},
        "calculation_type": "separator",
        "eos_type": "peng_robinson",
        "separator_config": {
            "reservoir_pressure_pa": 3.0e7, "reservoir_temperature_k": 380.0,
            "include_stock_tank": True,
            "separator_stages": [
                {"pressure_pa": 3.0e6, "temperature_k": 320.0, "name": "HP"},
                {"pressure_pa": 5.0e5, "temperature_k": 300.0, "name": "LP"},
            ],
        },
    })


def _separator_res() -> RunResult:
    return _completed(
        _separator_cfg(),
        separator_result=SeparatorResult(
            bo=1.21, rs=145.0, rs_scf_stb=815.0, bg=0.0042, api_gravity=39.5,
            stock_tank_oil_density=790.0, stock_tank_oil_mw_g_per_mol=174.25,
            stock_tank_oil_specific_gravity=0.8254,
            total_gas_moles=0.470000, shrinkage=0.8264,
            converged=True,
            stages=[
                SeparatorStageResult(
                    stage_number=1, stage_name="HP", pressure_pa=3.0e6,
                    temperature_k=320.0, vapor_fraction=0.35,
                    liquid_moles=0.65, vapor_moles=0.35,
                    liquid_density_kg_per_m3=640.0, vapor_density_kg_per_m3=85.0,
                    liquid_z_factor=0.24, vapor_z_factor=0.91, converged=True,
                ),
                SeparatorStageResult(
                    stage_number=2, stage_name="LP", pressure_pa=5.0e5,
                    temperature_k=300.0, vapor_fraction=0.18,
                    liquid_moles=0.53, vapor_moles=0.12,
                    liquid_density_kg_per_m3=710.0, vapor_density_kg_per_m3=22.0,
                    liquid_z_factor=0.07, vapor_z_factor=0.97, converged=True,
                ),
            ],
        ),
    )


def _stability_cfg() -> RunConfig:
    return _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.5},
            {"component_id": "C10", "mole_fraction": 0.5},
        ]},
        "calculation_type": "stability_analysis",
        "eos_type": "peng_robinson",
        "stability_analysis_config": {
            "pressure_pa": 10342135.5,
            "temperature_k": 326.76111111111106,
            "feed_phase": "liquid", "use_gdem": True,
            "n_random_trials": 2, "random_seed": 123,
            "max_eos_failures_per_trial": 4,
            "pressure_unit": "psia", "temperature_unit": "F",
        },
    })


def _stability_res() -> RunResult:
    config = _stability_cfg()
    vapor_trial = StabilityTrialResultData(
        kind="vapor_like", trial_phase="vapor",
        composition={"C1": 0.96, "C10": 0.04},
        tpd=-2.5e-2, iterations=6, total_iterations=10,
        converged=True, early_exit_unstable=True, n_phi_calls=18,
        n_eos_failures=1, message="negative tpd located",
        best_seed_index=0, candidate_seed_labels=["wilson", "extreme_lightest"],
        diagnostic_messages=["transient eos failure recovered"],
        seed_results=[
            StabilitySeedResultData(
                kind="vapor_like", trial_phase="vapor",
                seed_index=0, seed_label="wilson",
                initial_composition={"C1": 0.90, "C10": 0.10},
                composition={"C1": 0.96, "C10": 0.04},
                tpd=-2.5e-2, iterations=6, converged=True,
                early_exit_unstable=True, n_phi_calls=12,
                n_eos_failures=1, message="best seed",
            ),
            StabilitySeedResultData(
                kind="vapor_like", trial_phase="vapor",
                seed_index=1, seed_label="extreme_lightest",
                initial_composition={"C1": 0.97, "C10": 0.03},
                composition={"C1": 0.98, "C10": 0.02},
                tpd=-2.0e-2, iterations=4, converged=True,
                early_exit_unstable=False, n_phi_calls=6,
                n_eos_failures=0, message="alternate stationary point",
            ),
        ],
    )
    liquid_trial = StabilityTrialResultData(
        kind="liquid_like", trial_phase="liquid",
        composition={"C1": 0.22, "C10": 0.78},
        tpd=-8.0e-3, iterations=5, total_iterations=5,
        converged=True, early_exit_unstable=False, n_phi_calls=9,
        n_eos_failures=0, message=None,
        best_seed_index=0, candidate_seed_labels=["wilson"],
        diagnostic_messages=[],
        seed_results=[
            StabilitySeedResultData(
                kind="liquid_like", trial_phase="liquid",
                seed_index=0, seed_label="wilson",
                initial_composition={"C1": 0.30, "C10": 0.70},
                composition={"C1": 0.22, "C10": 0.78},
                tpd=-8.0e-3, iterations=5, converged=True,
                early_exit_unstable=False, n_phi_calls=9,
                n_eos_failures=0, message=None,
            )
        ],
    )
    return _completed(
        config,
        stability_analysis_result=StabilityAnalysisResult(
            stable=False, tpd_min=-2.5e-2,
            pressure_pa=config.stability_analysis_config.pressure_pa,
            temperature_k=config.stability_analysis_config.temperature_k,
            requested_feed_phase=config.stability_analysis_config.feed_phase,
            resolved_feed_phase="liquid", reference_root_used="liquid",
            phase_regime="two_phase", physical_state_hint="two_phase",
            physical_state_hint_basis="two_phase_regime",
            physical_state_hint_confidence="high",
            liquid_root_z=0.120000, vapor_root_z=0.910000,
            root_gap=7.900000e-01, gibbs_gap=2.500000e-01,
            average_reduced_pressure=1.230000,
            feed_composition={"C1": 0.5, "C10": 0.5},
            best_unstable_trial_kind="vapor_like",
            vapor_like_trial=vapor_trial, liquid_like_trial=liquid_trial,
        ),
    )


def _cancelled_res() -> RunResult:
    return RunResult(
        run_id="cancelled-result", run_name="cancelled-result",
        status=RunStatus.CANCELLED,
        error_message="Calculation was cancelled by user",
        started_at=_ts(), completed_at=_ts(), duration_seconds=0.5,
        config=_pt_flash_cfg(),
    )


CONFIG_BUILDERS: tuple[Callable[[], RunConfig], ...] = (
    _pt_flash_cfg, _bubble_cfg, _dew_cfg, _envelope_cfg,
    _tbp_cfg, _cce_cfg, _dl_cfg, _cvd_cfg, _separator_cfg,
)
CONFIG_IDS: tuple[str, ...] = tuple(
    b().calculation_type.value for b in CONFIG_BUILDERS
)

RESULT_BUILDERS: tuple[tuple[str, Callable[[], RunResult], str, str], ...] = (
    ("pt_flash", _pt_flash_res, "Component", "Pt Flash"),
    ("bubble_point", _bubble_res, "Component", "Bubble point"),
    ("dew_point", _dew_res, "Component", "Dew point"),
    ("phase_envelope", _envelope_res, "Type", "Phase envelope"),
    # First-column header text is now the compact "P ({unit})" form
    # across every pressure-step calc type (CCE / DL / CVD) after the
    # DL compaction pass — previously DL used "Pressure ({unit})" which
    # truncated in the right-rail, and CVD was hardcoded to bar.
    ("cce", _cce_res, "P (psia)", "CCE"),
    ("dl", _dl_res, "P (psia)", "Differential liberation"),
    ("cvd", _cvd_res, "P (psia)", "CVD"),
    ("separator", _separator_res, "Stage", "Separator train"),
)


def _summary_values(widget: ResultsTableWidget) -> dict[str, str]:
    return {
        widget.summary_table.item(row, 0).text(): widget.summary_table.item(row, 1).text()
        for row in range(widget.summary_table.rowCount())
    }


# Roundtrip tolerance: GUI spinboxes display pressures at 2 decimals in
# psia and temperatures at 2 decimals in °F, so a Pa→psia→Pa (or K→°F→K)
# round-trip through the UI can drift by half the last-displayed digit
# times the conversion factor:
#   2-dp psia × 6894.757 Pa/psia  ≈ ±34 Pa → use ±50 Pa for headroom.
#   2-dp °F × (5/9) K/°F          ≈ ±0.003 K → use ±0.01 K for headroom.
# These are defined once so every roundtrip assertion uses the same
# justified tolerance rather than the default rel=1e-6 which is tighter
# than the widget precision itself.
_PRESSURE_ROUNDTRIP_ABS_PA = 50.0
_TEMPERATURE_ROUNDTRIP_ABS_K = 0.01


def _assert_configs_equivalent(actual: RunConfig, expected: RunConfig) -> None:
    assert actual.calculation_type == expected.calculation_type
    assert actual.eos_type == expected.eos_type
    if expected.composition is None:
        assert actual.composition is None
    else:
        assert actual.composition is not None
        assert actual.composition.model_dump(mode="json") == expected.composition.model_dump(mode="json")
    assert actual.solver_settings.model_dump(mode="json") == expected.solver_settings.model_dump(mode="json")

    if expected.pt_flash_config is not None:
        assert actual.pt_flash_config is not None
        assert actual.pt_flash_config.pressure_pa == pytest.approx(expected.pt_flash_config.pressure_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.pt_flash_config.temperature_k == pytest.approx(expected.pt_flash_config.temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.pt_flash_config.pressure_unit == expected.pt_flash_config.pressure_unit
        assert actual.pt_flash_config.temperature_unit == expected.pt_flash_config.temperature_unit
    elif expected.bubble_point_config is not None:
        assert actual.bubble_point_config is not None
        assert actual.bubble_point_config.temperature_k == pytest.approx(expected.bubble_point_config.temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.bubble_point_config.pressure_initial_pa == pytest.approx(expected.bubble_point_config.pressure_initial_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.bubble_point_config.pressure_unit == expected.bubble_point_config.pressure_unit
        assert actual.bubble_point_config.temperature_unit == expected.bubble_point_config.temperature_unit
    elif expected.dew_point_config is not None:
        assert actual.dew_point_config is not None
        assert actual.dew_point_config.temperature_k == pytest.approx(expected.dew_point_config.temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.dew_point_config.pressure_initial_pa == pytest.approx(expected.dew_point_config.pressure_initial_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.dew_point_config.pressure_unit == expected.dew_point_config.pressure_unit
        assert actual.dew_point_config.temperature_unit == expected.dew_point_config.temperature_unit
    elif expected.phase_envelope_config is not None:
        assert actual.phase_envelope_config is not None
        assert actual.phase_envelope_config.temperature_min_k == pytest.approx(expected.phase_envelope_config.temperature_min_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.phase_envelope_config.temperature_max_k == pytest.approx(expected.phase_envelope_config.temperature_max_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.phase_envelope_config.n_points == expected.phase_envelope_config.n_points
        assert actual.phase_envelope_config.tracing_method == expected.phase_envelope_config.tracing_method
    elif expected.tbp_config is not None:
        assert actual.tbp_config is not None
        assert actual.tbp_config.cut_start == expected.tbp_config.cut_start
        assert [c.model_dump(mode="json") for c in actual.tbp_config.cuts] == [
            c.model_dump(mode="json") for c in expected.tbp_config.cuts
        ]
    elif expected.cce_config is not None:
        assert actual.cce_config is not None
        assert actual.cce_config.temperature_k == pytest.approx(expected.cce_config.temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.cce_config.pressure_start_pa == pytest.approx(expected.cce_config.pressure_start_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.cce_config.pressure_end_pa == pytest.approx(expected.cce_config.pressure_end_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.cce_config.n_steps == expected.cce_config.n_steps
        assert actual.cce_config.pressure_unit == expected.cce_config.pressure_unit
        assert actual.cce_config.temperature_unit == expected.cce_config.temperature_unit
    elif expected.dl_config is not None:
        assert actual.dl_config is not None
        assert actual.dl_config.temperature_k == pytest.approx(expected.dl_config.temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.dl_config.bubble_pressure_pa == pytest.approx(expected.dl_config.bubble_pressure_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.dl_config.pressure_end_pa == pytest.approx(expected.dl_config.pressure_end_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.dl_config.n_steps == expected.dl_config.n_steps
        assert actual.dl_config.pressure_unit == expected.dl_config.pressure_unit
        assert actual.dl_config.temperature_unit == expected.dl_config.temperature_unit
    elif expected.cvd_config is not None:
        assert actual.cvd_config is not None
        assert actual.cvd_config.temperature_k == pytest.approx(expected.cvd_config.temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.cvd_config.dew_pressure_pa == pytest.approx(expected.cvd_config.dew_pressure_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.cvd_config.pressure_end_pa == pytest.approx(expected.cvd_config.pressure_end_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.cvd_config.n_steps == expected.cvd_config.n_steps
    elif expected.separator_config is not None:
        assert actual.separator_config is not None
        assert actual.separator_config.reservoir_pressure_pa == pytest.approx(expected.separator_config.reservoir_pressure_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.separator_config.reservoir_temperature_k == pytest.approx(expected.separator_config.reservoir_temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
        assert actual.separator_config.include_stock_tank == expected.separator_config.include_stock_tank
        assert len(actual.separator_config.separator_stages) == len(expected.separator_config.separator_stages)
        for a_stage, e_stage in zip(actual.separator_config.separator_stages, expected.separator_config.separator_stages, strict=True):
            assert a_stage.name == e_stage.name
            assert a_stage.pressure_pa == pytest.approx(e_stage.pressure_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
            assert a_stage.temperature_k == pytest.approx(e_stage.temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
    elif expected.stability_analysis_config is not None:
        assert actual.stability_analysis_config is not None
        assert actual.stability_analysis_config.pressure_pa == pytest.approx(expected.stability_analysis_config.pressure_pa, abs=_PRESSURE_ROUNDTRIP_ABS_PA)
        assert actual.stability_analysis_config.temperature_k == pytest.approx(expected.stability_analysis_config.temperature_k, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)
    else:
        raise AssertionError("Expected config did not include a calculation-specific payload")


# ═══════════════════════════════════════════════════════════════════════════
#  1. test_main_window_opens
# ═══════════════════════════════════════════════════════════════════════════

def test_main_window_opens(window: PVTSimulatorWindow) -> None:
    assert window.composition_widget is not None
    assert window.conditions_widget is not None
    assert window.results_table is not None
    assert window.results_plot is not None
    assert window.workspace is not None
    assert window.workspace.results_pane is not None
    assert window.workspace.results_pane.minimumWidth() == 420
    assert window.workspace.results_pane.maximumWidth() == 420
    assert window.results_plot._view_mode == "generic"
    assert window.results_sidebar.layout().count() == 1
    assert window.results_sidebar.layout().itemAt(0).widget() is window.results_sidebar.table_widget
    assert window.unit_converter_widget is not None
    # Unit converter defaults: psia → atm. atm is more common in petroleum
    # courses than bar, so it's the target side by default. The source side
    # stays psia to match the repo-wide US-petroleum pressure default.
    # 1 psia = 6894.757 Pa; 1 atm = 101325 Pa → 1 psia ≈ 0.068046 atm.
    assert window.unit_converter_widget.result_value.text() == "0.068046 atm"


# ═══════════════════════════════════════════════════════════════════════════
#  2. test_config_roundtrip_all_calc_types
# ═══════════════════════════════════════════════════════════════════════════

@pytest.mark.parametrize("builder", CONFIG_BUILDERS, ids=CONFIG_IDS)
def test_config_roundtrip_all_calc_types(
    window: PVTSimulatorWindow,
    builder: Callable[[], RunConfig],
) -> None:
    config = builder()
    if config.composition is not None:
        window.composition_widget.set_composition(config.composition)
    window.conditions_widget.load_from_run_config(config)
    rebuilt = window._build_config()
    assert rebuilt is not None
    _assert_configs_equivalent(rebuilt, config)


@pytest.mark.parametrize(
    "label,builder",
    [
        ("plus_fraction_bubble", _bubble_plus_cfg),
        ("plus_fraction_tbp_fit", _pt_flash_plus_tbp_fit_cfg),
        ("inline_pseudo_bubble", _inline_pseudo_bubble_cfg),
        ("stability", _stability_cfg),
    ],
)
def test_config_roundtrip_advanced_compositions(
    window: PVTSimulatorWindow,
    label: str,
    builder: Callable[[], RunConfig],
) -> None:
    config = builder()
    if config.composition is not None:
        window.composition_widget.set_composition(config.composition)
    window.conditions_widget.load_from_run_config(config)
    rebuilt = window._build_config()
    assert rebuilt is not None
    _assert_configs_equivalent(rebuilt, config)


def test_config_roundtrip_saved_run_inputs(
    window: PVTSimulatorWindow, tmp_path: Path,
) -> None:
    config = _bubble_plus_cfg()
    run_dir = tmp_path / "saved-bubble"
    run_dir.mkdir()
    with (run_dir / "config.json").open("w", encoding="utf-8") as f:
        json.dump(config.model_dump(mode="json"), f, indent=2)
    window._load_saved_run_inputs(str(run_dir))
    rebuilt = window._build_config()
    assert rebuilt is not None
    _assert_configs_equivalent(rebuilt, config)
    assert window.status_label.text() == "Loaded inputs: saved-bubble"


def test_config_roundtrip_exact_schedules(window: PVTSimulatorWindow) -> None:
    cce = _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.70},
            {"component_id": "C4", "mole_fraction": 0.20},
            {"component_id": "C10", "mole_fraction": 0.10},
        ]},
        "calculation_type": "cce", "eos_type": "peng_robinson",
        "cce_config": {"temperature_k": 360.0, "pressure_points_pa": [2.0e7, 1.3e7, 2.0e6]},
    })
    window.composition_widget.set_composition(cce.composition)
    window.conditions_widget.load_from_run_config(cce)
    rebuilt = window._build_config()
    assert rebuilt is not None
    assert rebuilt.cce_config.pressure_points_pa == pytest.approx([2.0e7, 1.3e7, 2.0e6])

    dl = _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.40},
            {"component_id": "C3", "mole_fraction": 0.30},
            {"component_id": "C10", "mole_fraction": 0.30},
        ]},
        "calculation_type": "differential_liberation", "eos_type": "peng_robinson",
        "dl_config": {
            "temperature_k": 350.0, "bubble_pressure_pa": 1.5e7,
            "pressure_points_pa": [5.0e6, 3.0e6, 1.0e6],
        },
    })
    window.composition_widget.set_composition(dl.composition)
    window.conditions_widget.load_from_run_config(dl)
    rebuilt = window._build_config()
    assert rebuilt is not None
    assert rebuilt.dl_config.pressure_points_pa == pytest.approx([5.0e6, 3.0e6, 1.0e6])


def test_conditions_widget_builds_all_config_types(app: QApplication) -> None:
    # All pressure / temperature widgets now default to US-petroleum
    # units (psia, °F). The conversions used below:
    #   380 K = 106.85 °C = 224.33 °F
    #   350 K = 76.85 °C  = 170.33 °F
    #   3.0e7 Pa  = 300 bar  = 4351.13 psia
    #   1.5e7 Pa  = 150 bar  = 2175.57 psia
    #   1.0e6 Pa  =   10 bar = 145.04  psia
    #   5.652e6 Pa ~= 56.52 bar ~= 819.77 psia  (CVD dew default)
    #   5.0e6 Pa  =   50 bar = 725.19  psia     (CVD end default)
    # Separator stage defaults match the widget's new starting table
    # (HP ≈ 500 psia / 120 °F, LP ≈ 100 psia / 80 °F).
    w = ConditionsInputWidget()
    w.set_calculation_type(CalculationType.CVD)
    w.cvd_temperature.setValue(224.33)   # 380 K in °F
    w.cvd_p_dew.setValue(819.77)         # 5.652e6 Pa in psia
    w.cvd_p_end.setValue(725.19)         # 5.0e6 Pa in psia
    w.cvd_n_steps.setValue(12)
    cvd = w.get_cvd_config()
    assert cvd is not None
    assert cvd.temperature_k == pytest.approx(380.0, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)

    w.set_calculation_type(CalculationType.BUBBLE_POINT)
    w.bubble_temperature.setText("170.33")                 # 350 K in °F
    w.bubble_pressure_guess_enabled.setChecked(True)
    w.bubble_pressure_guess.setText("1813.06")             # 125e5 Pa in psia
    bp = w.get_bubble_point_config()
    assert bp is not None
    assert bp.temperature_k == pytest.approx(350.0, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)

    w.set_calculation_type(CalculationType.DEW_POINT)
    w.dew_temperature.setText("224.33")                    # 380 K in °F
    w.dew_pressure_guess_enabled.setChecked(True)
    w.dew_pressure_guess.setText("3045.95")                # 2.1e7 Pa in psia
    dp = w.get_dew_point_config()
    assert dp is not None
    assert dp.temperature_k == pytest.approx(380.0, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)

    w.set_calculation_type(CalculationType.SEPARATOR)
    w.separator_reservoir_pressure.setValue(4351.13)       # 3.0e7 Pa in psia
    w.separator_reservoir_temperature.setValue(224.33)     # 380 K in °F
    w.separator_include_stock_tank.setChecked(False)
    w._set_separator_stage_rows([
        {"name": "HP", "pressure_psia": 435.11, "temperature_f": 116.33},  # 30 bar / 46.85 °C
        {"name": "LP", "pressure_psia": 72.52,  "temperature_f": 80.33},   # 5  bar / 26.85 °C
    ])
    sep = w.get_separator_config()
    assert sep is not None
    assert sep.reservoir_pressure_pa == pytest.approx(3.0e7, abs=_PRESSURE_ROUNDTRIP_ABS_PA)

    w.set_calculation_type(CalculationType.DL)
    w.dl_temperature.setValue(170.33)                      # 350 K in °F
    w.set_dl_bubble_pressure_pa(1.5e7)
    w.dl_p_end.setValue(145.04)                            # 1.0e6 Pa in psia
    w.dl_n_steps.setValue(8)
    dl = w.get_dl_config()
    assert dl is not None
    assert dl.temperature_k == pytest.approx(350.0, abs=_TEMPERATURE_ROUNDTRIP_ABS_K)

    w.set_calculation_type(CalculationType.TBP)
    w.tbp_cut_start_spin.setValue(7)
    w._set_tbp_cut_rows([
        {"name": "C7-C9", "z": 0.020, "mw": 103.0, "sg": 0.74, "tb_k": 385.0},
        {"name": "C12", "z": 0.015, "mw": 170.0, "sg": 0.77},
        {"name": "C15-C18", "z": 0.015, "mw": 235.0, "sg": 0.80},
    ])
    tbp = w.get_tbp_config()
    assert tbp is not None
    assert [c.name for c in tbp.cuts] == ["C7-C9", "C12", "C15-C18"]

    combo_types = [CalculationType(w.calc_type_combo.itemData(i)) for i in range(w.calc_type_combo.count())]
    assert combo_types == list(GUI_SUPPORTED_CALCULATION_TYPES)
    eos_types = [EOSType(w.eos_combo.itemData(i)) for i in range(w.eos_combo.count())]
    assert eos_types == list(GUI_SUPPORTED_EOS_TYPES)


# ═══════════════════════════════════════════════════════════════════════════
#  3. test_results_render_all_calc_types
# ═══════════════════════════════════════════════════════════════════════════

@pytest.mark.parametrize(
    ("name", "builder", "expected_header", "expected_text"),
    RESULT_BUILDERS,
)
def test_results_render_all_calc_types(
    app: QApplication, name: str,
    builder: Callable[[], RunResult],
    expected_header: str, expected_text: str,
) -> None:
    result = builder()
    table = ResultsTableWidget()
    table.display_result(result)
    assert table.summary_table.rowCount() > 0
    assert table.composition_table.rowCount() > 0
    assert table.composition_table.horizontalHeaderItem(0).text() == expected_header

    plot = ResultsPlotWidget()
    if not getattr(plot, "_matplotlib_available", False):
        pytest.skip("matplotlib Qt backend unavailable")
    plot.display_result(result)
    assert len(plot.figure.axes) >= 1

    text = TextOutputWidget()
    text.display_result(result)
    assert expected_text in text.text.toPlainText()


def test_tbp_results_display_sections(app: QApplication) -> None:
    widget = ResultsTableWidget()
    widget.display_result(_tbp_res())
    summary = _summary_values(widget)
    assert summary["Cut Start"] == "7"
    assert summary["MW+"] == "164.800 g/mol"
    assert summary["Bridge Status"] == "Characterized SCN"
    assert widget.composition_section.title() == "Cuts"
    assert widget.details_section.title() == "Curves"

    text = TextOutputWidget()
    text.display_result(_tbp_res())
    report = text.text.toPlainText()
    assert "TBP assay" in report
    assert "Runtime bridge" in report
    assert "Characterized SCN" in report


def test_cvd_results_display_sections(app: QApplication) -> None:
    result = _cvd_res()
    table = ResultsTableWidget()
    table.display_result(result)
    assert table.composition_table.horizontalHeaderItem(2).text() == "Gas Produced"
    assert table.details_table.horizontalHeaderItem(1).text() == "Liquid Density"

    plot = ResultsPlotWidget()
    if not getattr(plot, "_matplotlib_available", False):
        pytest.skip("matplotlib Qt backend unavailable")
    plot.display_result(result)
    assert {a.text() for a in plot.series_menu.actions()} >= {
        "Liquid Dropout", "Gas Produced", "Cumulative Gas", "Z-factor",
        "Liquid Viscosity", "Vapor Viscosity",
    }


def test_cce_results_surface_density_columns_and_plot(app: QApplication) -> None:
    result = _cce_res()
    table = ResultsTableWidget()
    table.resize(420, 900)
    table.display_result(result)
    table.show()
    app.processEvents()
    # The CCE "Phase Properties" section was split into separate
    # "Phase Densities" and "Phase Viscosities" tables so all columns fit
    # the right rail without a horizontal scrollbar. ``details_section``
    # now points at the densities table; viscosities live on ``viscosity_section``.
    assert table.details_section.title() == "Phase Densities"
    assert table.details_table.horizontalHeaderItem(1).text() == "Liquid Density"
    assert table.details_table.item(0, 1).text() == "648.20"

    text = TextOutputWidget()
    text.display_result(result)
    report = text.text.toPlainText()
    # The CCE text-report subscripts are rendered as plain ASCII L/V
    # suffixes on the Greek densities/viscosities (``ρL`` not ``rhoL``,
    # ``ρV`` not ``rhoV``) because Consolas draws subscript glyphs at a
    # narrower visual width than ASCII digits, which would drift the
    # right-edge alignment of every header column.
    assert "\u03c1L" in report and "648.20" in report

    plot = ResultsPlotWidget()
    if not getattr(plot, "_matplotlib_available", False):
        pytest.skip("matplotlib Qt backend unavailable")
    plot.display_result(result)
    ax = plot.figure.axes[0]
    assert ax.get_ylabel() == "Density (kg/m³)"
    assert plot.series_controls.isHidden() is False


def test_separator_results_surface_sections(app: QApplication) -> None:
    result = _separator_res()
    table = ResultsTableWidget()
    table.display_result(result)
    summary = _summary_values(table)
    assert summary["Stock-tank MW"] == "174.250 g/mol"
    assert summary["Shrinkage"] == "0.8264"
    assert table.details_table.item(0, 1).text() == "640.00"


def test_dl_results_surface_sections(app: QApplication) -> None:
    result = _dl_res()
    table = ResultsTableWidget()
    table.display_result(result)
    summary = _summary_values(table)
    assert summary["Residual Oil Density"] == "782.00 kg/m³"
    assert table.composition_table.horizontalHeaderItem(4).text() == "Bg"
    # DL Details header compacted to standard petroleum-engineering symbols
    # (Greek letter + single-capital phase suffix) so all 8 columns fit the
    # right-rail width without truncation. Column 2 is now ρO (rho + O).
    assert table.details_table.horizontalHeaderItem(2).text() == "\u03c1O"


def test_stability_result_widgets_surface_trial_diagnostics(app: QApplication) -> None:
    result = _stability_res()
    table = ResultsTableWidget()
    table.display_result(result)
    summary = _summary_values(table)
    assert summary["Stable"] == "No"
    assert summary["Minimum TPD"] == "-2.500000e-02"
    assert summary["Phase Regime"] == "Two Phase"
    assert table.composition_table.horizontalHeaderItem(2).text() == "Vapor-like"

    text = TextOutputWidget()
    text.display_result(result)
    report = text.text.toPlainText()
    assert "Stability analysis" in report
    assert "Minimum TPD = -2.500000e-02" in report
    assert "wilson" in report

    diag = DiagnosticsWidget()
    diag.display_result(result)
    assert diag.status_label.text() == "UNSTABLE"
    assert diag.iterations_label.text() == "15"

    plot = ResultsPlotWidget()
    if not getattr(plot, "_matplotlib_available", False):
        pytest.skip("matplotlib Qt backend unavailable")
    plot.display_result(result)
    assert plot.series_controls.isHidden() is True
    assert plot.figure.axes[0].get_ylabel() == "TPD"


def test_pt_flash_widgets_display_density_viscosity_and_ift(app: QApplication) -> None:
    result = run_calculation(RunConfig.model_validate({
        "run_name": "PT Flash - viscosity wiring",
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.5},
            {"component_id": "C10", "mole_fraction": 0.5},
        ]},
        "calculation_type": "pt_flash", "eos_type": "peng_robinson",
        "pt_flash_config": {"pressure_pa": 5.0e6, "temperature_k": 350.0},
    }), write_artifacts=False)
    assert result.status == RunStatus.COMPLETED
    flash = result.pt_flash_result
    assert flash is not None

    table = ResultsTableWidget()
    table.display_result(result)
    summary = _summary_values(table)
    assert summary["Liquid Density"] == f"{flash.liquid_density_kg_per_m3:.2f} kg/m³"
    assert summary["Interfacial Tension"] == f"{flash.interfacial_tension_mn_per_m:.4f} mN/m"

    text = TextOutputWidget()
    text.display_result(result)
    assert f"Liquid density: {flash.liquid_density_kg_per_m3:.2f} kg/m³" in text.text.toPlainText()


@pytest.mark.xfail(
    reason=(
        "Parked until the phase-envelope module gets its accuracy/continuity "
        "redesign (tracked in the demo-prep handoff). The current plotter "
        "intentionally does not inject the detected critical point into the "
        "bubble or dew polylines — per the comment in results_view.py "
        "_plot_phase_envelope._curve_xy, the critical point generally does "
        "not lie on the discrete traced locus and sorting by T would create "
        "fake segments (spikes). This test asserts the opposite (connected "
        "curves through the critical point) and so is kept xfail as a "
        "reminder that the intended behavior depends on what the redesign "
        "decides — either the plot injects the CP cleanly (and this test "
        "passes) or we delete the test along with the redesign."
    ),
    strict=False,
)
def test_phase_envelope_plot_connects_curves_through_critical_point(app: QApplication) -> None:
    result = _envelope_res()
    plot = ResultsPlotWidget()
    if not getattr(plot, "_matplotlib_available", False):
        pytest.skip("matplotlib Qt backend unavailable")
    plot.display_result(result)
    ax = plot.figure.axes[0]
    bubble_line = next(l for l in ax.lines if l.get_label() == "Bubble Point")
    dew_line = next(l for l in ax.lines if l.get_label() == "Dew Point")
    cp = result.phase_envelope_result.critical_point
    crit_t = cp.temperature_k - 273.15
    crit_p = cp.pressure_pa / 1e5
    assert bubble_line.get_xdata()[-1] == pytest.approx(crit_t)
    assert bubble_line.get_ydata()[-1] == pytest.approx(crit_p)
    assert dew_line.get_xdata()[0] == pytest.approx(crit_t)
    assert dew_line.get_ydata()[0] == pytest.approx(crit_p)


def test_results_table_captures_and_exports(app: QApplication, tmp_path: Path) -> None:
    table = ResultsTableWidget()
    table.display_result(_bubble_res())
    table.capture_current_summary()
    table.display_result(_cce_res())
    table.capture_current_summary()
    assert table.captured_table.rowCount() == 2

    csv_path = tmp_path / "captured.csv"
    table._export_captured_csv(str(csv_path))
    with csv_path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 2
    assert rows[0]["Calculation"] == "Bubble Point"


def test_cce_plot_separates_incompatible_series(app: QApplication) -> None:
    plot = ResultsPlotWidget()
    if not getattr(plot, "_matplotlib_available", False):
        pytest.skip("matplotlib Qt backend unavailable")
    plot.display_result(_cce_res())
    for action in plot.series_menu.actions():
        if action.text() in {"Liquid Fraction", "Relative Volume", "Liquid Density"}:
            action.setChecked(True)
    plot._refresh_plot()
    assert len(plot.figure.axes) >= 3


def test_results_table_display_units(app: QApplication) -> None:
    config = _run_config({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.70},
            {"component_id": "C4", "mole_fraction": 0.20},
            {"component_id": "C10", "mole_fraction": 0.10},
        ]},
        "calculation_type": "cce", "eos_type": "peng_robinson",
        "cce_config": {
            "temperature_k": 360.0,
            "pressure_points_pa": [10342135.5, 8618446.25],
            "pressure_unit": "psia", "temperature_unit": "F",
        },
    })
    result = _completed(config, cce_result=CCEResult(
        temperature_k=360.0, saturation_pressure_pa=9652659.8,
        steps=[
            CCEStepResult(pressure_pa=10342135.5, relative_volume=1.00, liquid_fraction=1.0, vapor_fraction=0.0, z_factor=0.82, liquid_density_kg_per_m3=648.2),
            CCEStepResult(pressure_pa=8618446.25, relative_volume=1.15, liquid_fraction=0.72, vapor_fraction=0.28, z_factor=0.91, liquid_density_kg_per_m3=583.4, vapor_density_kg_per_m3=128.6),
        ],
    ))
    table = ResultsTableWidget()
    table.display_result(result)
    summary = _summary_values(table)
    assert summary["Temperature"] == "188.33 °F"
    assert summary["Saturation Pressure"] == "1400.00 psia"
    # Composition-table first column is rendered as ``f"P ({unit})"`` (the
    # compact ``P`` prefix is what's actually drawn, not the word ``Pressure``).
    assert table.composition_table.horizontalHeaderItem(0).text() == "P (psia)"


# ═══════════════════════════════════════════════════════════════════════════
#  4. test_composition_widget_validation
# ═══════════════════════════════════════════════════════════════════════════

def test_composition_widget_validation(app: QApplication) -> None:
    w = CompositionInputWidget()

    w.table.setRowCount(0)
    w._add_component_row("C1", 0.7)
    w._add_component_row("C7", 0.3)
    w._add_component_row("C1", 0.0)
    is_valid, msg = w.validate()
    assert is_valid is True
    comp = w.get_composition()
    assert [e.component_id for e in comp.components] == ["C1", "C7"]

    w.table.setRowCount(0)
    w._add_component_row("C1", 0.5)
    w._add_component_row("C1", 0.5)
    is_valid, msg = w.validate()
    assert is_valid is False
    assert "Duplicate component IDs" in msg

    w.table.setRowCount(0)
    w._add_component_row("C4", 0.5)
    w._add_component_row("nC4", 0.5)
    is_valid, msg = w.validate()
    assert is_valid is False
    assert "alias resolution" in msg


@pytest.mark.parametrize("mode", ["plus", "inline"])
def test_composition_widget_heavy_fraction_modes(app: QApplication, mode: str) -> None:
    w = CompositionInputWidget()
    w.table.setRowCount(0)

    if mode == "plus":
        w._add_component_row("C1", 0.7)
        w.set_calculation_type_context(CalculationType.BUBBLE_POINT)
        w.heavy_mode.setCurrentIndex(w.heavy_mode.findData(HEAVY_MODE_PLUS))
        w.plus_z_edit.setText("0.3")
        w.plus_mw_edit.setText("150.0")
        w.plus_sg_edit.setText("0.82")
        is_valid, msg = w.validate()
        assert is_valid is True
        comp = w.get_composition()
        assert comp.plus_fraction is not None
        assert comp.plus_fraction.z_plus == pytest.approx(0.3)
    else:
        w._add_component_row("C1", 0.6)
        w._add_component_row("", 0.4)
        combo = w.table.cellWidget(1, 0)
        combo.setCurrentText("PSEUDO+")
        w.inline_name_edit.setText("PSEUDO+")
        w.inline_mw_edit.setText("150.0")
        w.inline_tc_edit.setText("520.0")
        w.inline_pc_edit.setText("3500000.0")
        w.inline_omega_edit.setText("0.45")
        is_valid, msg = w.validate()
        assert is_valid is True
        comp = w.get_composition()
        assert [e.component_id for e in comp.components] == ["C1", "PSEUDO_PLUS"]
        assert len(comp.inline_components) == 1


def test_composition_normalization(app: QApplication) -> None:
    w = CompositionInputWidget()
    w.table.setRowCount(0)
    w._add_component_row("C1", 0.8)
    w._add_component_row("", 0.4)
    combo = w.table.cellWidget(1, 0)
    combo.setCurrentText("C7+")
    w.plus_mw_edit.setText("150.0")
    w.plus_sg_edit.setText("0.82")
    w._normalize()
    comp = w.get_composition()
    assert comp.components[0].mole_fraction == pytest.approx(0.666667, abs=1e-6)
    assert comp.plus_fraction.z_plus == pytest.approx(0.333333, abs=1e-6)


def test_plus_fraction_advanced_characterization_roundtrip(app: QApplication) -> None:
    w = CompositionInputWidget()
    w.table.setRowCount(0)
    w._add_component_row("C1", 0.7333)
    w.heavy_mode.setCurrentIndex(w.heavy_mode.findData(HEAVY_MODE_PLUS))
    w.plus_characterization_preset.setCurrentIndex(
        w.plus_characterization_preset.findData(PlusFractionCharacterizationPreset.MANUAL)
    )
    w.plus_z_edit.setText("0.2667")
    w.plus_mw_edit.setText("119.7876")
    w.plus_sg_edit.setText("0.82")
    w.plus_end_spin.setValue(20)
    w.plus_split_method.setCurrentText("lohrenz")
    w.plus_lumping_enabled.setChecked(True)
    w.plus_lumping_groups_spin.setValue(6)
    comp = w.get_composition()
    assert comp.plus_fraction.split_method == "lohrenz"

    reloaded = CompositionInputWidget()
    reloaded.set_composition(comp)
    assert reloaded.plus_split_method.currentText() == "lohrenz"
    assert reloaded.plus_lumping_groups_spin.value() == 6


def test_plus_fraction_tbp_fit_controls_roundtrip(app: QApplication) -> None:
    w = CompositionInputWidget()
    w.table.setRowCount(0)
    w._add_component_row("C1", 0.95)
    w.heavy_mode.setCurrentIndex(w.heavy_mode.findData(HEAVY_MODE_PLUS))
    w.plus_characterization_preset.setCurrentIndex(
        w.plus_characterization_preset.findData(PlusFractionCharacterizationPreset.MANUAL)
    )
    w.plus_cut_start_spin.setValue(7)
    w.plus_z_edit.setText("0.05")
    w.plus_sg_edit.setText("0.82")
    w.plus_split_method.setCurrentText("pedersen")
    w.plus_pedersen_solve_ab_from.setCurrentIndex(
        w.plus_pedersen_solve_ab_from.findData("fit_to_tbp")
    )
    w._set_plus_tbp_cut_rows([
        {"name": "C7", "z": 0.020, "mw": 96.0},
        {"name": "C8", "z": 0.015, "mw": 110.0},
        {"name": "C9", "z": 0.015, "mw": 124.0, "tb_k": 425.0},
    ])
    comp = w.get_composition()
    assert comp.plus_fraction.pedersen_solve_ab_from == "fit_to_tbp"
    assert len(comp.plus_fraction.tbp_cuts) == 3

    reloaded = CompositionInputWidget()
    reloaded.set_composition(comp)
    assert str(reloaded.plus_pedersen_solve_ab_from.currentData()) == "fit_to_tbp"
    assert reloaded.plus_tbp_cut_table.rowCount() == 3


def test_component_table_layout(app: QApplication) -> None:
    w = CompositionInputWidget()
    w.table.setRowCount(0)
    w._sync_table_height()
    for cid in ["C1", "C2", "C3", "nC4", "nC5"]:
        w._add_component_row(cid, 0.0)
    w.show()
    app.processEvents()
    initial_height = w.table.height()
    w._add_component_row("C6", 0.0)
    app.processEvents()
    assert w.table.rowCount() == 6
    assert w.table.verticalScrollBarPolicy() == Qt.ScrollBarPolicy.ScrollBarAlwaysOff
    assert w.table.height() > initial_height

    assert w.table.objectName() == "CompositionInputTable"
    assert w.heavy_tabs.documentMode() is True
    options = [w.table.cellWidget(0, 0).itemText(i) for i in range(w.table.cellWidget(0, 0).count())]
    assert options == list(COMPONENT_PICKER_OPTIONS)


def test_standard_component_picker_only_lists_resolvable_components() -> None:
    assert STANDARD_COMPONENTS is not None
    for cid in STANDARD_COMPONENTS:
        assert resolve_component_id(cid)


# ═══════════════════════════════════════════════════════════════════════════
#  5. test_gui_agrees_with_backend
# ═══════════════════════════════════════════════════════════════════════════

def test_gui_agrees_with_backend(app: QApplication) -> None:
    config = RunConfig.model_validate({
        "composition": {"components": [
            {"component_id": "C1", "mole_fraction": 0.55},
            {"component_id": "C10", "mole_fraction": 0.45},
        ]},
        "calculation_type": "pt_flash", "eos_type": "peng_robinson",
        "pt_flash_config": {"pressure_pa": 8.0e6, "temperature_k": 350.0},
    })
    result = run_calculation(config, write_artifacts=False)
    assert result.status == RunStatus.COMPLETED
    flash = result.pt_flash_result
    assert flash is not None

    table = ResultsTableWidget()
    table.display_result(result)
    summary = _summary_values(table)
    # PTFlashConfig now defaults to the repo-wide US-petroleum units
    # (PSIA, °F), matching every other calc type after the consistency
    # pass. The summary and report both pick up the new defaults:
    #   Pa→psia conversion uses factor 6894.757 (schemas.py PSIA mapping),
    #   so 8e6 / 6894.757 = 1160.30195 psia (5 decimals) / 1160.30 (2 decimals).
    #   350 K = 76.85 °C = 170.33 °F exactly (76.85 × 1.8 + 32).
    assert summary["Pressure"] == "1160.30 psia"
    assert summary["Temperature"] == "170.33 °F"
    assert summary["Vapor Fraction"] == f"{flash.vapor_fraction:.6f}"

    text = TextOutputWidget()
    text.display_result(result)
    report = text.text.toPlainText()
    assert "T = 170.330 °F" in report
    assert "P = 1160.30195 psia" in report


# ═══════════════════════════════════════════════════════════════════════════
#  6. test_zoom_and_scaling
# ═══════════════════════════════════════════════════════════════════════════

def test_zoom_and_scaling(window: PVTSimulatorWindow) -> None:
    initial_scale = window.ui_scale
    initial_fixed_min_width = window.workspace.fixed_pane.minimumWidth()
    initial_results_min_width = window.workspace.results_pane.minimumWidth()
    initial_text_font_size = window.text_output_widget.text.font().pointSizeF()

    window._zoom_in()
    zoomed = DEFAULT_UI_SCALE + UI_SCALE_STEP
    assert window.ui_scale == pytest.approx(zoomed)
    assert window.workspace.fixed_pane.minimumWidth() == scale_metric(
        initial_fixed_min_width, zoomed, reference_scale=DEFAULT_UI_SCALE,
    )
    assert window.workspace.results_pane.minimumWidth() == scale_metric(
        initial_results_min_width, zoomed, reference_scale=DEFAULT_UI_SCALE,
    )
    assert window.text_output_widget.text.font().pointSizeF() > initial_text_font_size

    window._reset_zoom()
    assert window.ui_scale == pytest.approx(initial_scale)
    assert window.workspace.fixed_pane.minimumWidth() == initial_fixed_min_width


def test_results_table_scales_with_zoom(app: QApplication) -> None:
    table = ResultsTableWidget()
    table.resize(340, 900)
    table.display_result(_bubble_res())
    table.show()
    app.processEvents()
    initial_width = table.summary_table.columnWidth(0)
    table.apply_ui_scale(DEFAULT_UI_SCALE + UI_SCALE_STEP)
    app.processEvents()
    assert table.summary_table.columnWidth(0) > initial_width


def test_conditions_widget_scales_with_zoom(app: QApplication) -> None:
    w = ConditionsInputWidget()
    initial_width = w.pressure_unit.maximumWidth()
    w.apply_ui_scale(DEFAULT_UI_SCALE + UI_SCALE_STEP)
    assert w.pressure_unit.maximumWidth() > initial_width


def test_plot_surface_resizes_canvas(app: QApplication) -> None:
    plot = ResultsPlotWidget()
    if not getattr(plot, "_matplotlib_available", False):
        pytest.skip("matplotlib Qt backend unavailable")
    plot.resize(720, 540)
    plot.show()
    plot.display_result(_cce_res())
    app.processEvents()
    initial_size = plot.canvas.size()
    plot.resize(1180, 860)
    app.processEvents()
    assert plot.canvas.width() > initial_size.width()


def test_results_tables_fit_inside_right_rail(
    window: PVTSimulatorWindow, app: QApplication,
) -> None:
    window.resize(1800, 1000)
    window.show()
    window.results_sidebar.display_result(_inline_pseudo_bubble_res())
    app.processEvents()
    table = window.results_table

    def _col_span(grid):
        return sum(grid.columnWidth(c) for c in range(grid.columnCount()))

    for grid in (table.summary_table, table.composition_table, table.details_table):
        assert grid.verticalHeader().isVisible() is False
        assert _col_span(grid) <= grid.viewport().width() + 1


# ═══════════════════════════════════════════════════════════════════════════
#  7. test_dark_theme_palette
# ═══════════════════════════════════════════════════════════════════════════

def test_dark_theme_palette(app: QApplication) -> None:
    plot = ResultsPlotWidget()
    if not getattr(plot, "_matplotlib_available", False):
        pytest.skip("matplotlib Qt backend unavailable")
    assert plot.palette().color(plot.backgroundRole()).name().lower() == QColor(PLOT_SURFACE_COLOR).name().lower()

    plot.display_result(_envelope_res())
    assert QColor(PLOT_CANVAS_COLOR).name().lower() == QColor(PLOT_SURFACE_COLOR).name().lower()
    assert plot.figure.get_facecolor()[:3] == pytest.approx(QColor(PLOT_CANVAS_COLOR).getRgbF()[:3], abs=1e-3)
    ax = plot.figure.axes[0]
    assert ax.get_facecolor()[:3] == pytest.approx(QColor(PLOT_CANVAS_COLOR).getRgbF()[:3], abs=1e-3)

    stylesheet = build_cato_stylesheet(scale=DEFAULT_UI_SCALE)
    assert "QLabel" in stylesheet
    assert "background: transparent;" in stylesheet
    assert "QComboBox::drop-down" in stylesheet
    assert "QGroupBox" in stylesheet


# ═══════════════════════════════════════════════════════════════════════════
#  8. test_workspace_layout
# ═══════════════════════════════════════════════════════════════════════════

def test_workspace_layout(app: QApplication) -> None:
    fixed_inputs = QLabel("inputs")
    fixed_results = QLabel("results")
    view_specs = [
        ViewSpec("critical_props", "Critical prop."),
        ViewSpec("phase_envelope", "Phase Envelope"),
        ViewSpec("text_output", "Text output"),
    ]
    view_widgets = {
        "critical_props": QWidget(),
        "phase_envelope": QWidget(),
        "text_output": QWidget(),
    }

    ws = TwoPaneWorkspace(
        view_specs=view_specs, view_widgets=view_widgets,
        left_default="text_output", right_default="phase_envelope",
        fixed_widget=fixed_inputs, fixed_title="Feeds / Inputs",
        fixed_width=360, fixed_right_widget=fixed_results,
        fixed_right_title="Results table", fixed_right_width=340,
        default_pane_mode="single",
    )

    assert ws.fixed_pane is not None
    assert ws.results_pane is not None
    assert isinstance(ws.outer_splitter, QSplitter)
    assert ws.fixed_pane.minimumWidth() == 360
    assert ws.fixed_pane.maximumWidth() == 360
    assert ws.results_pane.minimumWidth() == 340
    assert ws.results_pane.maximumWidth() == 340
    assert ws.outer_splitter.count() == 3
    assert ws.outer_splitter.handleWidth() == 0
    assert ws.pane_mode == "single"
    assert ws.right_pane.isHidden()


def test_workspace_scale_resizes(app: QApplication) -> None:
    ws = TwoPaneWorkspace(
        view_specs=[ViewSpec("a", "A"), ViewSpec("b", "B")],
        view_widgets={"a": QWidget(), "b": QWidget()},
        left_default="a", right_default="b",
        fixed_widget=QLabel("inputs"), fixed_title="Inputs", fixed_width=360,
        fixed_right_widget=QLabel("results"), fixed_right_title="Results", fixed_right_width=340,
    )
    initial = ws.fixed_pane.minimumWidth()
    scaled = DEFAULT_UI_SCALE + 0.2
    ws.apply_ui_scale(scaled, previous_scale=DEFAULT_UI_SCALE)
    assert ws.fixed_pane.minimumWidth() == scale_metric(initial, scaled, reference_scale=DEFAULT_UI_SCALE)


def test_workspace_toggle_panes(app: QApplication) -> None:
    ws = TwoPaneWorkspace(
        view_specs=[ViewSpec("a", "A"), ViewSpec("b", "B"), ViewSpec("c", "C")],
        view_widgets={"a": QWidget(), "b": QWidget(), "c": QWidget()},
        left_default="a", right_default="b",
        fixed_widget=QLabel("inputs"),
        fixed_right_widget=QLabel("results"),
        default_pane_mode="single",
    )
    ws.set_pane_mode("double")
    assert ws.pane_mode == "double"
    assert not ws.right_pane.isHidden()
    assert ws.panes_splitter.handleWidth() > 0


# ═══════════════════════════════════════════════════════════════════════════
#  9. test_run_log_lifecycle
# ═══════════════════════════════════════════════════════════════════════════

def test_run_log_lifecycle(
    app: QApplication, monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    run_a = str(Path("C:/tmp/run-a"))
    run_b = str(Path("C:/tmp/run-b"))
    run_c = str(Path("C:/tmp/run-c"))
    run_results = {
        run_a: _completed(
            _pt_flash_cfg().model_copy(update={"run_name": "alpha"}),
            pt_flash_result=_pt_flash_res().pt_flash_result,
        ).model_copy(update={"run_name": "alpha"}),
        run_c: _completed(
            _pt_flash_cfg().model_copy(update={"run_name": "charlie"}),
            pt_flash_result=_pt_flash_res().pt_flash_result,
        ).model_copy(update={"run_name": "charlie"}),
        run_b: _completed(
            _pt_flash_cfg().model_copy(update={"run_name": "bravo"}),
            pt_flash_result=_pt_flash_res().pt_flash_result,
        ).model_copy(update={"run_name": "bravo"}),
    }
    runs = [{"path": run_a}, {"path": run_c}, {"path": run_b}]

    monkeypatch.setattr("pvtapp.widgets.run_log_view.list_runs", lambda limit=200: runs[:limit])
    monkeypatch.setattr(
        "pvtapp.widgets.run_log_view.load_run_result",
        lambda run_dir: run_results.get(str(run_dir)),
    )

    widget = RunLogWidget()
    widget.show()
    app.processEvents()

    assert widget.tree.topLevelItemCount() == 3
    assert widget.tree.header().sortIndicatorSection() == 2

    widget._on_header_clicked(0)
    app.processEvents()
    names = [widget.tree.topLevelItem(i).text(0) for i in range(widget.tree.topLevelItemCount())]
    assert names == ["alpha", "bravo", "charlie"]

    widget.preview_plot.display_result = lambda _r: None  # type: ignore[method-assign]
    item = widget.tree.topLevelItem(0)
    widget._on_item_clicked(item, 0)
    app.processEvents()

    widget._set_preview(None, None)
    app.processEvents()


def test_run_log_grouping(
    app: QApplication, monkeypatch: pytest.MonkeyPatch,
) -> None:
    run_a = str(Path("C:/tmp/run-a"))
    run_b = str(Path("C:/tmp/run-b"))
    run_c = str(Path("C:/tmp/run-c"))
    run_results = {
        run_a: _completed(
            _bubble_cfg().model_copy(update={"run_name": "charlie"}),
            bubble_point_result=_bubble_res().bubble_point_result,
        ).model_copy(update={"run_name": "charlie"}),
        run_b: _completed(
            _bubble_cfg().model_copy(update={"run_name": "alpha"}),
            bubble_point_result=_bubble_res().bubble_point_result,
        ).model_copy(update={"run_name": "alpha"}),
        run_c: _completed(
            _pt_flash_cfg().model_copy(update={"run_name": "bravo"}),
            pt_flash_result=_pt_flash_res().pt_flash_result,
        ).model_copy(update={"run_name": "bravo"}),
    }
    runs = [{"path": run_a}, {"path": run_b}, {"path": run_c}]

    monkeypatch.setattr("pvtapp.widgets.run_log_view.list_runs", lambda limit=200: runs[:limit])
    monkeypatch.setattr(
        "pvtapp.widgets.run_log_view.load_run_result",
        lambda run_dir: run_results.get(str(run_dir)),
    )

    widget = RunLogWidget()
    widget.show()
    app.processEvents()
    widget.group_by_combo.setCurrentText("Test Type")
    widget._on_header_clicked(0)
    app.processEvents()

    assert widget.tree.topLevelItemCount() == 2
    bubble_group = widget.tree.topLevelItem(0)
    assert bubble_group.text(0) == "Bubble Point"
    assert bubble_group.childCount() == 2
    assert bubble_group.child(0).text(0) == "alpha"


def test_run_log_delete_and_export(
    app: QApplication, monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    run_a = tmp_path / "run-a"
    run_b = tmp_path / "run-b"
    run_a.mkdir()
    run_b.mkdir()
    (run_a / "config.json").write_text('{"run":"a"}', encoding="utf-8")
    (run_a / "results.json").write_text('{"result":"a"}', encoding="utf-8")
    (run_b / "config.json").write_text('{"run":"b"}', encoding="utf-8")
    (run_b / "results.json").write_text('{"result":"b"}', encoding="utf-8")
    run_results = {
        str(run_a): _pt_flash_res(),
        str(run_b): _bubble_res(),
    }
    archive_path = tmp_path / "selected-runs.zip"

    monkeypatch.setattr(
        "pvtapp.widgets.run_log_view.list_runs",
        lambda limit=200: [{"path": str(run_a)}, {"path": str(run_b)}],
    )
    monkeypatch.setattr(
        "pvtapp.widgets.run_log_view.load_run_result",
        lambda run_path: run_results.get(str(run_path)),
    )
    monkeypatch.setattr(
        "pvtapp.widgets.run_log_view.QFileDialog.getSaveFileName",
        lambda *args, **kwargs: (str(archive_path), "ZIP Files (*.zip)"),
    )
    monkeypatch.setattr(
        "pvtapp.widgets.run_log_view.QMessageBox.question",
        lambda *args, **kwargs: QMessageBox.StandardButton.Yes,
    )

    widget = RunLogWidget()
    widget.preview_plot.display_result = lambda _r: None  # type: ignore[method-assign]
    widget.show()
    app.processEvents()

    item_a = widget.tree.topLevelItem(0)
    item_b = widget.tree.topLevelItem(1)
    widget.tree.setCurrentItem(item_b)
    item_a.setSelected(True)
    item_b.setSelected(True)
    widget._on_item_clicked(item_b, 0)

    widget._export_selected()
    assert archive_path.exists()
    with zipfile.ZipFile(archive_path) as archive:
        names = set(archive.namelist())
    assert "run-a/config.json" in names
    assert "run-b/results.json" in names

    widget._delete_selected()
    assert run_a.exists() is False
    assert run_b.exists() is False


def test_run_log_cached_results_sidebar(
    window: PVTSimulatorWindow,
    monkeypatch: pytest.MonkeyPatch,
    app: QApplication,
) -> None:
    run_dir = "C:/tmp/run-a"
    expected_run_dir = str(Path(run_dir))
    result = _inline_pseudo_bubble_res()

    monkeypatch.setattr("pvtapp.widgets.run_log_view.list_runs", lambda limit=200: [{"path": run_dir}])
    monkeypatch.setattr(
        "pvtapp.widgets.run_log_view.load_run_result",
        lambda run_path: result if str(run_path) == expected_run_dir else None,
    )
    monkeypatch.setattr(window.run_log_widget, "refresh", lambda: None)

    window._on_calculation_finished(_bubble_res())
    assert window.workspace.results_pane._title_label.text() == "Bubble Point Results"

    window.run_log_widget.refresh = RunLogWidget.refresh.__get__(window.run_log_widget)
    window.run_log_widget.refresh()
    app.processEvents()
    assert window.run_log_widget.tree.topLevelItemCount() == 1
    item = window.run_log_widget.tree.topLevelItem(0)
    window.run_log_widget._on_item_clicked(item, 0)
    app.processEvents()
    assert window.results_table.display_is_cached is True


def test_main_window_restores_persisted_zoom(
    app: QApplication, monkeypatch: pytest.MonkeyPatch, settings_path: Path,
) -> None:
    def _create_settings(_self) -> QSettings:
        return QSettings(str(settings_path), QSettings.Format.IniFormat)

    monkeypatch.setattr(PVTSimulatorWindow, "_create_settings", _create_settings)

    first = PVTSimulatorWindow()
    try:
        first._zoom_in()
        first._zoom_in()
        assert first.ui_scale == pytest.approx(DEFAULT_UI_SCALE + (2 * UI_SCALE_STEP))
    finally:
        first.close()
        first.deleteLater()
        app.processEvents()
        if QCoreApplication is not None and QEvent is not None:
            QCoreApplication.sendPostedEvents(None, QEvent.Type.DeferredDelete.value)
            app.processEvents()
        gc.collect()

    second = PVTSimulatorWindow()
    try:
        assert second.ui_scale == pytest.approx(DEFAULT_UI_SCALE + (2 * UI_SCALE_STEP))
    finally:
        second.close()
        second.deleteLater()
        app.processEvents()
        if QCoreApplication is not None and QEvent is not None:
            QCoreApplication.sendPostedEvents(None, QEvent.Type.DeferredDelete.value)
            app.processEvents()
        gc.collect()


@pytest.mark.parametrize(
    ("name", "builder", "expected_header", "expected_cell"),
    (
        ("pt_flash", _pt_flash_res, "Component", "C1"),
        ("bubble_point", _bubble_res, "Component", "C1"),
        ("dew_point", _dew_res, "Component", "C1"),
        ("phase_envelope", _envelope_res, "Type", "bubble"),
        # CSV columns mirror the new US-petroleum defaults (psia, °F).
        # Expected numeric fragments are the psia-converted equivalents
        # of the fixtures' pressure_pa values:
        #   CCE  pressure_pa = 2.0e7 → 2900.75 psia (fixture)
        #   DL   pressure_pa = 1.5e7 → 2175.57 psia (fixture)
        #   CVD  pressure_pa = 5.652e6 → 819.77 psia (fixture)
        ("cce", _cce_res, "Pressure_psia", "2900."),
        ("dl", _dl_res, "Pressure_psia", "2175."),
        ("cvd", _cvd_res, "Pressure_psia", "819."),
        ("separator", _separator_res, "Stage", "HP"),
    ),
)
def test_csv_export_for_supported_results(
    window: PVTSimulatorWindow, tmp_path: Path,
    name: str, builder: Callable[[], RunResult],
    expected_header: str, expected_cell: str,
) -> None:
    result = builder()
    filename = tmp_path / f"{name}.csv"
    window._export_csv(result, str(filename))
    assert filename.exists()
    with filename.open("r", newline="") as f:
        rows = list(csv.reader(f))
    assert rows[0][0] == expected_header
    assert any(expected_cell in cell for row in rows[1:] for cell in row)


def test_csv_export_rejects_cancelled_results(
    window: PVTSimulatorWindow, tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    filename = tmp_path / "cancelled.csv"
    warnings: list[tuple[str, str]] = []

    def fake_warning(_parent, title: str, message: str):
        warnings.append((title, message))
        return QMessageBox.StandardButton.Ok

    monkeypatch.setattr(QMessageBox, "warning", fake_warning)
    window._export_csv(_cancelled_res(), str(filename))
    assert warnings == [("Export Error", "CSV export is only available for completed calculations")]
    assert not filename.exists()


# ── Excel export ─────────────────────────────────────────────────────
# The Excel exporter builds a multi-sheet workbook per calc type —
# "Summary" plus one data sheet per logical section (Expansion / Phase
# Densities / Per-Step Compositions / etc.). The contract checks below
# assert the sheet structure is stable and at least one per-calc unit
# header is rendered correctly (psia / °F), so a regression in either
# the sheet layout or the unit labelling would be caught.


@pytest.mark.parametrize(
    ("name", "builder", "expected_sheets", "header_sheet", "header_needle"),
    (
        ("pt_flash", _pt_flash_res, {"Summary", "Composition"}, "Composition", "Feed z"),
        ("bubble_point", _bubble_res, {"Summary", "K-Values"}, "K-Values", "K-value"),
        ("dew_point", _dew_res, {"Summary", "K-Values"}, "K-Values", "K-value"),
        ("cce", _cce_res, {"Summary", "Expansion", "Phase Densities", "Phase Viscosities"}, "Expansion", "P (psia)"),
        ("dl", _dl_res, {"Summary", "Steps", "Phase Properties"}, "Steps", "P (psia)"),
        ("cvd", _cvd_res, {"Summary", "Steps", "Phase Densities", "Phase Viscosities"}, "Steps", "P (psia)"),
        ("separator", _separator_res, {"Summary", "Stages"}, "Stages", "T (\u00b0F)"),
        ("phase_envelope", _envelope_res, {"Summary", "Bubble Curve", "Dew Curve", "Critical Point"}, "Bubble Curve", "T (\u00b0F)"),
        ("tbp", _tbp_res, {"Summary", "Cuts"}, "Cuts", "Name"),
    ),
)
def test_excel_export_for_supported_results(
    window: PVTSimulatorWindow,
    tmp_path: Path,
    name: str,
    builder: Callable[[], RunResult],
    expected_sheets: set[str],
    header_sheet: str,
    header_needle: str,
) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    result = builder()
    filename = tmp_path / f"{name}.xlsx"
    window._export_excel(result, str(filename))
    assert filename.exists()

    wb = load_workbook(str(filename), read_only=True)
    assert expected_sheets.issubset(set(wb.sheetnames)), (
        f"Expected {expected_sheets} in {wb.sheetnames}"
    )
    headers = [cell.value for cell in next(wb[header_sheet].iter_rows(max_row=1))]
    assert header_needle in headers, (
        f"Expected header containing '{header_needle}' in {headers}"
    )


def test_excel_export_rejects_cancelled_results(
    window: PVTSimulatorWindow, tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    pytest.importorskip("openpyxl")
    filename = tmp_path / "cancelled.xlsx"
    warnings: list[tuple[str, str]] = []

    def fake_warning(_parent, title: str, message: str):
        warnings.append((title, message))
        return QMessageBox.StandardButton.Ok

    monkeypatch.setattr(QMessageBox, "warning", fake_warning)
    window._export_excel(_cancelled_res(), str(filename))
    assert warnings == [("Export Error", "Excel export is only available for completed calculations")]
    assert not filename.exists()


def test_conditions_widget_misc_behaviors(app: QApplication) -> None:
    w = ConditionsInputWidget()

    w.set_calculation_type(CalculationType.CVD)
    w.cvd_temperature.setValue(106.85)
    w.cvd_p_dew.setValue(40.0)
    w.cvd_p_end.setValue(50.0)
    assert w.get_cvd_config() is None

    w.set_calculation_type(CalculationType.DL)
    w.dl_temperature.setValue(76.85)
    w.dl_p_end.setValue(10.0)
    w.dl_n_steps.setValue(8)
    is_valid, _ = w.validate()
    assert is_valid is True
    assert w.dl_bubble_pressure.isReadOnly() is True

    combo = w.pressure_unit
    initial_index = combo.currentIndex()

    class DummyWheel:
        def __init__(self):
            self.ignored = False
        def ignore(self):
            self.ignored = True

    ev = DummyWheel()
    combo.wheelEvent(ev)
    assert ev.ignored is True
    assert combo.currentIndex() == initial_index

    assert w.solver_group.title() == "Tolerance / Solver Settings"
    assert w.solver_group.isCheckable() is False


def test_mole_fraction_editor_and_alignment(app: QApplication) -> None:
    w = CompositionInputWidget()
    w.table.setRowCount(0)
    w._add_component_row("C1", 0.123456)

    item = w.table.item(0, 1)
    alignment = item.textAlignment()
    assert alignment & Qt.AlignmentFlag.AlignLeft
    assert alignment & Qt.AlignmentFlag.AlignVCenter

    delegate = w.table.itemDelegateForColumn(1)
    index = w.table.model().index(0, 1)
    option = QStyleOptionViewItem()
    option.font = w.table.font()
    editor = delegate.createEditor(w.table.viewport(), option, index)
    assert isinstance(editor, QLineEdit)
    assert "padding: 0px" in editor.styleSheet()
    assert "background: transparent" not in editor.styleSheet()


def test_component_dropdown_button_style(app: QApplication) -> None:
    w = CompositionInputWidget()
    w.table.setRowCount(0)
    w._add_component_row("C1", 0.5)
    combo = w.table.cellWidget(0, 0)
    assert combo.objectName() == "CompositionComponentCombo"
    assert f"width: {COMPONENT_DROPDOWN_BUTTON_WIDTH}px" in combo.styleSheet()


def test_heavy_fraction_tabs_layout(app: QApplication) -> None:
    w = CompositionInputWidget()
    assert w.heavy_tabs.usesScrollButtons() is False
    assert w.heavy_tabs.tabBar().expanding() is True
    assert w.heavy_tabs.tabBar().elideMode() == Qt.TextElideMode.ElideNone

    w.heavy_tabs.setCurrentIndex(1)

    class DummyWheel:
        def __init__(self):
            self.ignored = False
        def ignore(self):
            self.ignored = True

    ev = DummyWheel()
    w.heavy_tabs.tabBar().wheelEvent(ev)
    assert ev.ignored is True
    assert w.heavy_tabs.currentIndex() == 1


def test_enter_key_advances_fraction_row(app: QApplication) -> None:
    if QTest is None:
        pytest.skip("PySide6 QtTest is not installed")
    w = CompositionInputWidget()
    w.table.setRowCount(0)
    w._add_component_row("C1", 0.1)
    w._add_component_row("C2", 0.2)
    w.show()
    first = w.table.item(0, 1)
    w.table.setCurrentItem(first)
    w.table.editItem(first)
    app.processEvents()
    editor = w.table.findChild(QLineEdit)
    assert editor is not None
    editor.setText("0.300000")
    QTest.keyClick(editor, Qt.Key.Key_Return)
    app.processEvents()
    assert w.table.item(0, 1).text() == "0.300000"
    assert w.table.currentRow() == 1
    assert w.table.currentColumn() == 1
