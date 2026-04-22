"""Multi-sheet Excel (.xlsx) export for completed PVT run results.

Every pressure / temperature / density / viscosity / GOR / FVF value is
written in the repo's US-petroleum display units (psia, °F, ...) so the
workbook the user gets matches what they see on screen and in the CSV /
text-output exports. Each calc type produces a dedicated set of sheets:

* A "Summary" sheet with run metadata and per-calc highlights.
* One "data" sheet per logical section (e.g. CCE → ``Expansion`` /
  ``Phase Densities`` / ``Phase Viscosities`` / ``Per-Step Liquid`` /
  ``Per-Step Vapor``).

Data sheets use bold / filled header rows with frozen panes, numbers
are written as numbers (not strings) so the engineer can pivot / chart
without reformatting, and column headers carry their physical units in
parentheses (``P (psia)``, ``Liquid Density (kg/m³)``, ...).

The public entry point is :func:`export_result_to_excel`.
"""

from __future__ import annotations

from typing import Any, Callable, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from pvtapp.schemas import (
    PressureUnit,
    RunResult,
    TemperatureUnit,
    pressure_from_pa,
    temperature_from_k,
)


# A single counter used to mint unique Table display-names across every
# workbook export. Excel requires Table names to be workbook-unique, and
# using a running counter lets us mint "tbl_Expansion", "tbl_Phase_Densities",
# etc. without worrying about collisions between runs in the same process.
_TABLE_NAME_COUNTER: list[int] = [0]


def _next_table_name(hint: str) -> str:
    """Return a valid, unique Excel Table display name.

    Excel Table names must start with a letter, contain only letters /
    digits / underscores, and be unique within the workbook. The workbook
    is new on every export, so we only need global process-uniqueness to
    avoid surprises if the caller reuses an in-memory Workbook.
    """
    _TABLE_NAME_COUNTER[0] += 1
    sanitized = "".join(c if c.isalnum() else "_" for c in hint).strip("_")
    if not sanitized or not sanitized[0].isalpha():
        sanitized = "tbl_" + sanitized
    return f"{sanitized}_{_TABLE_NAME_COUNTER[0]}"


# ─────────────────────────────────────────────────────────────────────
# Styling constants. Match the GUI dark theme's cool-navy accents with
# a readable white-on-navy header band; data cells use openpyxl defaults
# (no fill, default font) so the workbook prints cleanly and remains
# easy to reformat downstream.
# ─────────────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_SUMMARY_SECTION_FONT = Font(bold=True, size=12, color="1E3A5F")
_SUMMARY_KEY_FONT = Font(bold=True)

# Built-in Excel Table style used for every data sheet. "TableStyleMedium2"
# is the stock "Blue, Table Style Medium 2" — the default users recognize
# from Ctrl+T in Excel. It gives banded alternate rows, a blue banner on
# the header, and AutoFilter dropdowns on every column for free.
_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

# Number formats keyed by column semantics. Attached to cells that hold
# floats so Excel renders a consistent precision without rounding the
# underlying stored value.
_FMT_PRESSURE = "0.00"
_FMT_TEMPERATURE = "0.00"
_FMT_DENSITY = "0.00"
_FMT_VISCOSITY = "0.0000"
_FMT_DIMENSIONLESS_LOW = "0.0000"   # e.g. fractions, Z-factor, Bo
_FMT_DIMENSIONLESS_HIGH = "0.000000" # e.g. mole fractions, moles remaining
_FMT_INTEGER = "0"


# ─────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────


def export_result_to_excel(result: RunResult, filename: str) -> None:
    """Write a multi-sheet workbook for ``result`` to ``filename``.

    Raises:
        RuntimeError: if the ``RunResult`` has no calculation-specific
            payload (cancelled / failed / unsupported calc type).
    """
    wb = Workbook()
    # Drop the default empty sheet so the first real sheet sits at index 0.
    default = wb.active
    if default is not None:
        wb.remove(default)

    meta_rows = _build_meta_rows(result)

    if result.pt_flash_result is not None:
        _write_pt_flash(wb, result, meta_rows)
    elif result.bubble_point_result is not None:
        _write_saturation(wb, result, meta_rows, kind="bubble")
    elif result.dew_point_result is not None:
        _write_saturation(wb, result, meta_rows, kind="dew")
    elif result.phase_envelope_result is not None:
        _write_phase_envelope(wb, result, meta_rows)
    elif result.cce_result is not None:
        _write_cce(wb, result, meta_rows)
    elif result.dl_result is not None:
        _write_dl(wb, result, meta_rows)
    elif result.cvd_result is not None:
        _write_cvd(wb, result, meta_rows)
    elif result.swelling_test_result is not None:
        _write_swelling(wb, result, meta_rows)
    elif result.separator_result is not None:
        _write_separator(wb, result, meta_rows)
    elif result.stability_analysis_result is not None:
        _write_stability(wb, result, meta_rows)
    elif result.tbp_result is not None:
        _write_tbp(wb, result, meta_rows)
    else:
        raise RuntimeError(
            "Result has no calculation-specific payload; nothing to export."
        )

    wb.save(filename)


# ─────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────


def _build_meta_rows(result: RunResult) -> list[tuple[str, Any]]:
    """Assemble the run-metadata rows shared across every calc type."""
    cfg = result.config
    duration = (
        float(result.duration_seconds) if result.duration_seconds is not None else None
    )
    completed = (
        result.completed_at.strftime("%Y-%m-%d %H:%M:%S")
        if result.completed_at is not None
        else ""
    )
    return [
        ("Run ID", result.run_id),
        ("Run Name", result.run_name or ""),
        ("Status", result.status.value),
        ("Calculation Type", cfg.calculation_type.value),
        ("EOS Type", cfg.eos_type.value),
        ("Completed At", completed),
        ("Duration (s)", duration),
    ]


def _display_composition(component_id: str) -> str:
    """Mirror the text-output helper: render PSEUDO_PLUS as PSEUDO+."""
    return "PSEUDO+" if component_id.strip() == "PSEUDO_PLUS" else component_id.strip()


def _write_summary_sheet(
    wb: Workbook,
    result_summary: list[tuple[str, Any]],
    *,
    meta_rows: Optional[list[tuple[str, Any]]] = None,
    title: str = "Summary",
) -> None:
    """Create the Summary sheet: two banded-header sub-tables.

    Matches the table-esque polish of the per-section data sheets —
    each sub-block ("Run Metadata" / "Result Summary") gets a navy
    "Property | Value" header band, bold property names, and values
    in the adjacent column. A blank row separates the two blocks.
    """
    ws = wb.create_sheet(title)
    row = 1

    def _write_block(block_title: str, block_rows: list[tuple[str, Any]]) -> None:
        nonlocal row
        # Section title above the block's own banner.
        title_cell = ws.cell(row=row, column=1, value=block_title)
        title_cell.font = _SUMMARY_SECTION_FONT
        ws.cell(row=row, column=2, value=None)
        row += 1
        # Navy-banded "Property | Value" header.
        for col_index, label in enumerate(("Property", "Value"), start=1):
            cell = ws.cell(row=row, column=col_index, value=label)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
        row += 1
        # Key-value rows.
        for prop, value in block_rows:
            ws.cell(row=row, column=1, value=prop).font = _SUMMARY_KEY_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1

    if meta_rows:
        _write_block("Run Metadata", meta_rows)
        row += 1  # blank separator
    _write_block("Result Summary", result_summary)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28
    ws.row_dimensions[1].height = 22  # breathing room above first banner


def _write_data_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: Iterable[list[Any]],
    *,
    number_formats: Optional[list[Optional[str]]] = None,
    column_widths: Optional[list[Optional[int]]] = None,
) -> Worksheet:
    """Create a data sheet with bold/filled headers and typed data rows.

    ``number_formats`` applies per-column to numeric cells; ``None`` in a
    slot leaves the cell's format at Excel's default ``General``. Column
    widths default to a safe auto-size (header length + padding, capped).
    The top row is frozen so scrolling keeps headers visible.
    """
    ws = wb.create_sheet(title)

    # Write header row first. Styling is applied by the Table style we
    # attach at the end — we don't need to manually fill/bold here
    # (doing so would fight the built-in table style).
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Materialize rows so we can count them (Table refs require a bounded
    # range) and still apply per-column number formats as we go.
    materialized_rows = list(rows)
    for r_index, row_values in enumerate(materialized_rows, start=2):
        for c_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r_index, column=c_index, value=value)
            if number_formats is not None:
                fmt = number_formats[c_index - 1] if c_index - 1 < len(number_formats) else None
                if fmt is not None and isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = fmt

    for col in range(1, len(headers) + 1):
        if column_widths is not None and col - 1 < len(column_widths) and column_widths[col - 1] is not None:
            ws.column_dimensions[get_column_letter(col)].width = column_widths[col - 1]
        else:
            # Auto-size from header length (padded), capped so long unit
            # strings don't push the layout out of shape.
            header_len = len(str(headers[col - 1]))
            ws.column_dimensions[get_column_letter(col)].width = max(12, min(header_len + 4, 28))

    ws.freeze_panes = "A2"
    # Wider header row so multi-line unit labels breathe.
    ws.row_dimensions[1].height = 28

    # Register this range as a real Excel Table. Gives the workbook
    # AutoFilter dropdowns on every column, banded alternate row
    # colouring, structured references (``=Table[@[P (psia)]]``), and
    # Ctrl+T-style sort/filter behaviour — all the things a user
    # expects when they open the file in Excel. Skip the Table if there
    # are no data rows (Excel requires at least one data row beyond the
    # header to wrap a valid Table reference).
    if materialized_rows:
        last_col_letter = get_column_letter(len(headers))
        last_row = len(materialized_rows) + 1  # +1 for header row
        table = Table(
            displayName=_next_table_name(title),
            ref=f"A1:{last_col_letter}{last_row}",
        )
        table.tableStyleInfo = _TABLE_STYLE
        ws.add_table(table)
    return ws


def _optional_number(value: Optional[float], *, require_positive: bool = False) -> Optional[float]:
    """Return ``value`` as a float, or ``None`` for sentinels / invalids."""
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if require_positive and f <= 0.0:
        return None
    return f


def _temperature_header(unit: TemperatureUnit) -> str:
    return f"T (\u00b0{unit.value})"


def _pressure_header(unit: PressureUnit) -> str:
    return f"P ({unit.value})"


def _per_step_composition_rows(
    steps: list,
    field: str,
    pressure_unit: PressureUnit,
) -> tuple[list[str], list[list[Any]]]:
    """Assemble headers + rows for a "Pressure × component" composition table.

    Steps whose ``field`` is missing or empty (e.g. absent phase in a
    single-phase step) are silently skipped — their presence as empty
    rows would just be noise in the sheet.
    """
    rows: list[list[Any]] = []
    component_ids: list[str] = []
    for step in steps:
        comp = getattr(step, field, None)
        if not comp:
            continue
        if not component_ids:
            component_ids = list(comp.keys())
        row: list[Any] = [pressure_from_pa(step.pressure_pa, pressure_unit)]
        for cid in component_ids:
            row.append(_optional_number(comp.get(cid)))
        rows.append(row)
    if not component_ids:
        return [_pressure_header(pressure_unit)], []
    headers = [_pressure_header(pressure_unit)] + [
        _display_composition(cid) for cid in component_ids
    ]
    return headers, rows


# ─────────────────────────────────────────────────────────────────────
# Per-calc-type writers
# ─────────────────────────────────────────────────────────────────────


def _write_pt_flash(wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]) -> None:
    r = result.pt_flash_result
    cfg = result.config.pt_flash_config
    p_unit = cfg.pressure_unit if cfg is not None else PressureUnit.PSIA
    t_unit = cfg.temperature_unit if cfg is not None else TemperatureUnit.F

    summary: list[tuple[str, Any]] = []
    if cfg is not None:
        summary.append(("Pressure", f"{pressure_from_pa(cfg.pressure_pa, p_unit):.2f} {p_unit.value}"))
        summary.append((
            "Temperature",
            f"{temperature_from_k(cfg.temperature_k, t_unit):.2f} \u00b0{t_unit.value}",
        ))
    summary.extend([
        ("Phase", r.phase),
        ("Vapor Fraction", float(r.vapor_fraction)),
        ("Liquid Density (kg/m³)", _optional_number(r.liquid_density_kg_per_m3)),
        ("Vapor Density (kg/m³)", _optional_number(r.vapor_density_kg_per_m3)),
        ("Liquid Viscosity (cP)", _optional_number(r.liquid_viscosity_cp)),
        ("Vapor Viscosity (cP)", _optional_number(r.vapor_viscosity_cp)),
        ("Interfacial Tension (mN/m)", _optional_number(r.interfacial_tension_mn_per_m)),
    ])
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    # Composition sheet: feed z, liquid x, vapor y, K — one row per component.
    feed_comp = result.config.composition.components if result.config.composition else []
    feed_map = {e.component_id: float(e.mole_fraction) for e in feed_comp}
    if result.config.composition and result.config.composition.plus_fraction is not None:
        pf = result.config.composition.plus_fraction
        feed_map[pf.label] = float(pf.z_plus)

    all_ids = sorted(
        set(feed_map) | set(r.display_liquid_composition) | set(r.display_vapor_composition)
    )
    rows = [
        [
            _display_composition(cid),
            _optional_number(feed_map.get(cid)),
            _optional_number(r.display_liquid_composition.get(cid)),
            _optional_number(r.display_vapor_composition.get(cid)),
            _optional_number(r.display_k_values.get(cid)),
        ]
        for cid in all_ids
    ]
    _write_data_sheet(
        wb,
        "Composition",
        ["Component", "Feed z", "Liquid x", "Vapor y", "K-value"],
        rows,
        number_formats=[None, _FMT_DIMENSIONLESS_HIGH, _FMT_DIMENSIONLESS_HIGH, _FMT_DIMENSIONLESS_HIGH, _FMT_DIMENSIONLESS_LOW],
    )


def _write_saturation(
    wb: Workbook,
    result: RunResult,
    meta_rows: list[tuple[str, Any]],
    *,
    kind: str,
) -> None:
    """Bubble- and dew-point results share the same layout."""
    r = result.bubble_point_result if kind == "bubble" else result.dew_point_result
    cfg = (
        result.config.bubble_point_config
        if kind == "bubble"
        else result.config.dew_point_config
    )
    p_unit = cfg.pressure_unit if cfg is not None else PressureUnit.PSIA
    t_unit = cfg.temperature_unit if cfg is not None else TemperatureUnit.F

    label = "Bubble Pressure" if kind == "bubble" else "Dew Pressure"
    summary: list[tuple[str, Any]] = [
        ("Converged", bool(r.converged)),
        ("Temperature", f"{temperature_from_k(r.temperature_k, t_unit):.2f} \u00b0{t_unit.value}"),
        (label, f"{pressure_from_pa(r.pressure_pa, p_unit):.2f} {p_unit.value}"),
        ("Iterations", int(r.iterations)),
        ("Residual", float(r.residual)),
    ]
    if kind == "bubble":
        summary.append(("Stable Liquid", bool(r.stable_liquid)))
    else:
        summary.append(("Stable Vapor", bool(getattr(r, "stable_vapor", False))))
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    all_ids = sorted(
        set(r.display_liquid_composition)
        | set(r.display_vapor_composition)
        | set(r.display_k_values)
    )
    rows = [
        [
            _display_composition(cid),
            _optional_number(r.display_liquid_composition.get(cid)),
            _optional_number(r.display_vapor_composition.get(cid)),
            _optional_number(r.display_k_values.get(cid)),
        ]
        for cid in all_ids
    ]
    _write_data_sheet(
        wb,
        "K-Values",
        ["Component", "Liquid x", "Vapor y", "K-value"],
        rows,
        number_formats=[None, _FMT_DIMENSIONLESS_HIGH, _FMT_DIMENSIONLESS_HIGH, _FMT_DIMENSIONLESS_LOW],
    )


def _write_phase_envelope(
    wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]
) -> None:
    r = result.phase_envelope_result
    # Phase envelope has no per-run unit preference on the config; use
    # the repo-wide US-petroleum defaults (psia, °F) so the workbook
    # matches every other calc type's Excel export.
    p_unit = PressureUnit.PSIA
    t_unit = TemperatureUnit.F

    summary: list[tuple[str, Any]] = [
        ("Bubble Points", len(r.bubble_curve)),
        ("Dew Points", len(r.dew_curve)),
    ]
    if r.critical_point is not None:
        summary.extend([
            (
                "Critical Temperature",
                f"{temperature_from_k(r.critical_point.temperature_k, t_unit):.2f} \u00b0{t_unit.value}",
            ),
            (
                "Critical Pressure",
                f"{pressure_from_pa(r.critical_point.pressure_pa, p_unit):.2f} {p_unit.value}",
            ),
        ])
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    def _curve_rows(points) -> list[list[Any]]:
        return [
            [
                temperature_from_k(pt.temperature_k, t_unit),
                pressure_from_pa(pt.pressure_pa, p_unit),
            ]
            for pt in points
        ]

    headers = [_temperature_header(t_unit), _pressure_header(p_unit)]
    formats = [_FMT_TEMPERATURE, _FMT_PRESSURE]

    if r.bubble_curve:
        _write_data_sheet(wb, "Bubble Curve", headers, _curve_rows(r.bubble_curve), number_formats=formats)
    if r.dew_curve:
        _write_data_sheet(wb, "Dew Curve", headers, _curve_rows(r.dew_curve), number_formats=formats)
    if r.critical_point is not None:
        _write_data_sheet(
            wb,
            "Critical Point",
            headers,
            [[
                temperature_from_k(r.critical_point.temperature_k, t_unit),
                pressure_from_pa(r.critical_point.pressure_pa, p_unit),
            ]],
            number_formats=formats,
        )


def _write_cce(wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]) -> None:
    r = result.cce_result
    cfg = result.config.cce_config
    p_unit = cfg.pressure_unit if cfg is not None else PressureUnit.PSIA
    t_unit = cfg.temperature_unit if cfg is not None else TemperatureUnit.F

    summary: list[tuple[str, Any]] = [
        ("Temperature", f"{temperature_from_k(r.temperature_k, t_unit):.2f} \u00b0{t_unit.value}"),
        ("Steps", len(r.steps)),
    ]
    if r.saturation_pressure_pa is not None:
        summary.append((
            "Saturation Pressure",
            f"{pressure_from_pa(r.saturation_pressure_pa, p_unit):.2f} {p_unit.value}",
        ))
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    # Expansion
    _write_data_sheet(
        wb,
        "Expansion",
        [
            _pressure_header(p_unit),
            "Rel. Vol. (V/Vsat)",
            "Liquid Frac.",
            "Vapor Frac.",
            "Z-factor",
        ],
        [
            [
                pressure_from_pa(step.pressure_pa, p_unit),
                float(step.relative_volume),
                _optional_number(step.liquid_fraction),
                _optional_number(step.vapor_fraction),
                _optional_number(step.z_factor),
            ]
            for step in r.steps
        ],
        number_formats=[_FMT_PRESSURE, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW],
    )

    # Phase Densities
    _write_data_sheet(
        wb,
        "Phase Densities",
        [_pressure_header(p_unit), "Liquid Density (kg/m³)", "Vapor Density (kg/m³)"],
        [
            [
                pressure_from_pa(step.pressure_pa, p_unit),
                _optional_number(step.liquid_density_kg_per_m3, require_positive=True),
                _optional_number(step.vapor_density_kg_per_m3, require_positive=True),
            ]
            for step in r.steps
        ],
        number_formats=[_FMT_PRESSURE, _FMT_DENSITY, _FMT_DENSITY],
    )

    # Phase Viscosities
    _write_data_sheet(
        wb,
        "Phase Viscosities",
        [_pressure_header(p_unit), "Liquid Viscosity (cP)", "Vapor Viscosity (cP)"],
        [
            [
                pressure_from_pa(step.pressure_pa, p_unit),
                _optional_number(step.liquid_viscosity_cp, require_positive=True),
                _optional_number(step.vapor_viscosity_cp, require_positive=True),
            ]
            for step in r.steps
        ],
        number_formats=[_FMT_PRESSURE, _FMT_VISCOSITY, _FMT_VISCOSITY],
    )

    # Per-step compositions
    for field, title in [
        ("liquid_composition", "Per-Step Liquid Composition"),
        ("vapor_composition", "Per-Step Vapor Composition"),
    ]:
        headers, rows = _per_step_composition_rows(r.steps, field, p_unit)
        if rows:
            formats = [_FMT_PRESSURE] + [_FMT_DIMENSIONLESS_HIGH] * (len(headers) - 1)
            _write_data_sheet(wb, title, headers, rows, number_formats=formats)


def _write_dl(wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]) -> None:
    r = result.dl_result
    cfg = result.config.dl_config
    p_unit = cfg.pressure_unit if cfg is not None else PressureUnit.PSIA
    t_unit = cfg.temperature_unit if cfg is not None else TemperatureUnit.F

    summary: list[tuple[str, Any]] = [
        ("Converged", bool(r.converged)),
        ("Temperature", f"{temperature_from_k(r.temperature_k, t_unit):.2f} \u00b0{t_unit.value}"),
        ("Bubble Pressure", f"{pressure_from_pa(r.bubble_pressure_pa, p_unit):.2f} {p_unit.value}"),
        ("Rsi (Initial Solution GOR)", float(r.rsi)),
        ("Boi (Initial FVF)", float(r.boi)),
        ("Residual Oil Density (kg/m³)", _optional_number(r.residual_oil_density_kg_per_m3)),
        ("Steps", len(r.steps)),
    ]
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    # Production properties per step.
    _write_data_sheet(
        wb,
        "Steps",
        [
            _pressure_header(p_unit),
            "RsD (Solution GOR)",
            "RsDb (Initial Rs at Pb)",
            "Bo",
            "Bg",
            "BtD (Total FVF)",
            "Vapor Frac.",
            "Cum. Gas",
            "Liquid Moles Remaining",
        ],
        [
            [
                pressure_from_pa(step.pressure_pa, p_unit),
                float(step.rs),
                # RsDb is the initial solution GOR at the bubble point
                # (constant for every DL step). Included as a per-row
                # column so the professor can plot RsD + RsDb vs pressure
                # directly from the exported deliverable — the PETE665
                # term project requires plotting both.
                float(r.rsi),
                float(step.bo),
                _optional_number(step.bg),
                float(step.bt),
                float(step.vapor_fraction),
                _optional_number(step.cumulative_gas_produced),
                _optional_number(step.liquid_moles_remaining),
            ]
            for step in r.steps
        ],
        number_formats=[
            _FMT_PRESSURE,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_HIGH,
        ],
    )

    # Phase properties per step.
    _write_data_sheet(
        wb,
        "Phase Properties",
        [
            _pressure_header(p_unit),
            "Oil Density (kg/m³)",
            "Oil Viscosity (cP)",
            "Gas Gravity",
            "Gas Z-factor",
            "Gas Viscosity (cP)",
        ],
        [
            [
                pressure_from_pa(step.pressure_pa, p_unit),
                _optional_number(step.oil_density_kg_per_m3, require_positive=True),
                _optional_number(step.oil_viscosity_cp, require_positive=True),
                _optional_number(step.gas_gravity),
                _optional_number(step.gas_z_factor),
                _optional_number(step.gas_viscosity_cp, require_positive=True),
            ]
            for step in r.steps
        ],
        number_formats=[_FMT_PRESSURE, _FMT_DENSITY, _FMT_VISCOSITY, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW, _FMT_VISCOSITY],
    )

    # Per-step compositions.
    for field, title in [
        ("liquid_composition", "Per-Step Liquid Composition"),
        ("gas_composition", "Per-Step Gas Composition"),
    ]:
        headers, rows = _per_step_composition_rows(r.steps, field, p_unit)
        if rows:
            formats = [_FMT_PRESSURE] + [_FMT_DIMENSIONLESS_HIGH] * (len(headers) - 1)
            _write_data_sheet(wb, title, headers, rows, number_formats=formats)


def _write_cvd(wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]) -> None:
    r = result.cvd_result
    # CVDConfig has no per-run unit preference; use repo-wide US-petroleum
    # defaults so this export matches the rest of the GUI.
    p_unit = PressureUnit.PSIA
    t_unit = TemperatureUnit.F

    summary: list[tuple[str, Any]] = [
        ("Converged", bool(r.converged)),
        ("Temperature", f"{temperature_from_k(r.temperature_k, t_unit):.2f} \u00b0{t_unit.value}"),
        ("Dew Pressure", f"{pressure_from_pa(r.dew_pressure_pa, p_unit):.2f} {p_unit.value}"),
        ("Initial Z", float(r.initial_z)),
        ("Steps", len(r.steps)),
    ]
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    _write_data_sheet(
        wb,
        "Steps",
        [
            _pressure_header(p_unit),
            "Liquid Dropout",
            "Gas Produced",
            "Cum. Gas",
            "Moles Remaining",
            "Z (2-phase)",
        ],
        [
            [
                pressure_from_pa(step.pressure_pa, p_unit),
                float(step.liquid_dropout),
                _optional_number(step.gas_produced),
                float(step.cumulative_gas_produced),
                _optional_number(step.moles_remaining),
                _optional_number(step.z_two_phase),
            ]
            for step in r.steps
        ],
        number_formats=[_FMT_PRESSURE, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_HIGH, _FMT_DIMENSIONLESS_LOW],
    )

    _write_data_sheet(
        wb,
        "Phase Densities",
        [_pressure_header(p_unit), "Liquid Density (kg/m³)", "Vapor Density (kg/m³)"],
        [
            [
                pressure_from_pa(step.pressure_pa, p_unit),
                _optional_number(step.liquid_density_kg_per_m3, require_positive=True),
                _optional_number(step.vapor_density_kg_per_m3, require_positive=True),
            ]
            for step in r.steps
        ],
        number_formats=[_FMT_PRESSURE, _FMT_DENSITY, _FMT_DENSITY],
    )
    _write_data_sheet(
        wb,
        "Phase Viscosities",
        [_pressure_header(p_unit), "Liquid Viscosity (cP)", "Vapor Viscosity (cP)"],
        [
            [
                pressure_from_pa(step.pressure_pa, p_unit),
                _optional_number(step.liquid_viscosity_cp, require_positive=True),
                _optional_number(step.vapor_viscosity_cp, require_positive=True),
            ]
            for step in r.steps
        ],
        number_formats=[_FMT_PRESSURE, _FMT_VISCOSITY, _FMT_VISCOSITY],
    )


def _write_swelling(
    wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]
) -> None:
    r = result.swelling_test_result
    cfg = result.config.swelling_test_config
    p_unit = cfg.pressure_unit if cfg is not None else PressureUnit.PSIA
    t_unit = cfg.temperature_unit if cfg is not None else TemperatureUnit.F

    summary: list[tuple[str, Any]] = [
        ("Overall Status", r.overall_status),
        ("Fully Certified", bool(r.fully_certified)),
        ("Temperature", f"{temperature_from_k(r.temperature_k, t_unit):.2f} \u00b0{t_unit.value}"),
    ]
    if r.baseline_bubble_pressure_pa is not None:
        summary.append((
            "Baseline Bubble Pressure",
            f"{pressure_from_pa(r.baseline_bubble_pressure_pa, p_unit):.2f} {p_unit.value}",
        ))
    summary.append(("Steps", len(r.steps)))
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    _write_data_sheet(
        wb,
        "Enrichment Steps",
        [
            "Added Gas (mol/mol oil)",
            "Total Mixture (mol/mol oil)",
            f"Bubble Pressure ({p_unit.value})",
            "Swelling Factor",
            "Sat. Liquid Density (kg/m³)",
            "Status",
            "Message",
        ],
        [
            [
                float(step.added_gas_moles_per_mole_oil),
                float(step.total_mixture_moles_per_mole_oil),
                None if step.bubble_pressure_pa is None else pressure_from_pa(step.bubble_pressure_pa, p_unit),
                _optional_number(step.swelling_factor),
                _optional_number(step.saturated_liquid_density_kg_per_m3, require_positive=True),
                step.status,
                step.message or "",
            ]
            for step in r.steps
        ],
        number_formats=[_FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW, _FMT_PRESSURE, _FMT_DIMENSIONLESS_LOW, _FMT_DENSITY, None, None],
        column_widths=[22, 24, 22, 18, 24, 22, 40],
    )


def _write_separator(
    wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]
) -> None:
    r = result.separator_result
    # SeparatorConfig has no per-run unit preference; use repo-wide defaults.
    p_unit = PressureUnit.PSIA
    t_unit = TemperatureUnit.F

    summary: list[tuple[str, Any]] = [
        ("Converged", bool(r.converged)),
        ("Bo", float(r.bo)),
        ("Rs (mol/mol)", float(r.rs)),
        ("Rs (scf/stb)", float(r.rs_scf_stb)),
        ("Bg", float(r.bg)),
        ("API Gravity", float(r.api_gravity)),
        ("Stock-Tank Oil Density (kg/m³)", float(r.stock_tank_oil_density)),
        ("Stock-Tank Oil MW (g/mol)", _optional_number(r.stock_tank_oil_mw_g_per_mol)),
        ("Stock-Tank Oil SG", _optional_number(r.stock_tank_oil_specific_gravity)),
        ("Total Gas Moles", _optional_number(r.total_gas_moles)),
        ("Shrinkage", _optional_number(r.shrinkage)),
        ("Stages", len(r.stages)),
    ]
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    _write_data_sheet(
        wb,
        "Stages",
        [
            "Stage",
            _pressure_header(p_unit),
            _temperature_header(t_unit),
            "Vapor Frac.",
            "Liquid Moles",
            "Vapor Moles",
            "Liquid Density (kg/m³)",
            "Vapor Density (kg/m³)",
            "Liquid Z",
            "Vapor Z",
            "Converged",
        ],
        [
            [
                stage.stage_name or f"Stage {stage.stage_number}",
                pressure_from_pa(stage.pressure_pa, p_unit),
                temperature_from_k(stage.temperature_k, t_unit),
                _optional_number(stage.vapor_fraction),
                _optional_number(stage.liquid_moles),
                _optional_number(stage.vapor_moles),
                _optional_number(stage.liquid_density_kg_per_m3, require_positive=True),
                _optional_number(stage.vapor_density_kg_per_m3, require_positive=True),
                _optional_number(stage.liquid_z_factor),
                _optional_number(stage.vapor_z_factor),
                bool(stage.converged),
            ]
            for stage in r.stages
        ],
        number_formats=[
            None,
            _FMT_PRESSURE,
            _FMT_TEMPERATURE,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_HIGH,
            _FMT_DIMENSIONLESS_HIGH,
            _FMT_DENSITY,
            _FMT_DENSITY,
            _FMT_DIMENSIONLESS_LOW,
            _FMT_DIMENSIONLESS_LOW,
            None,
        ],
    )


def _write_stability(
    wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]
) -> None:
    r = result.stability_analysis_result
    cfg = result.config.stability_analysis_config
    p_unit = cfg.pressure_unit if cfg is not None else PressureUnit.PSIA
    t_unit = cfg.temperature_unit if cfg is not None else TemperatureUnit.F

    summary: list[tuple[str, Any]] = [
        ("Stable", bool(r.stable)),
        ("Temperature", f"{temperature_from_k(r.temperature_k, t_unit):.2f} \u00b0{t_unit.value}"),
        ("Pressure", f"{pressure_from_pa(r.pressure_pa, p_unit):.2f} {p_unit.value}"),
        ("Minimum TPD", float(r.tpd_min)),
        ("Phase Regime", str(r.phase_regime)),
        ("Physical State Hint", str(r.physical_state_hint)),
        ("Hint Basis", str(r.physical_state_hint_basis)),
        ("Hint Confidence", str(r.physical_state_hint_confidence)),
    ]
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)


def _write_tbp(wb: Workbook, result: RunResult, meta_rows: list[tuple[str, Any]]) -> None:
    r = result.tbp_result

    summary: list[tuple[str, Any]] = [
        ("Cuts", len(r.cuts)),
    ]
    _write_summary_sheet(wb, summary, meta_rows=meta_rows)

    _write_data_sheet(
        wb,
        "Cuts",
        [
            "Name",
            "Carbon Number",
            "Mole Fraction (z)",
            "MW (g/mol)",
            "SG",
            "Boiling Point (K)",
            "Cum. Mole %",
            "Cum. Mass %",
        ],
        [
            [
                cut.name,
                int(cut.carbon_number),
                _optional_number(getattr(cut, "mole_fraction", None)),
                _optional_number(getattr(cut, "molecular_weight", None)),
                _optional_number(getattr(cut, "specific_gravity", None)),
                _optional_number(getattr(cut, "boiling_point_k", None)),
                _optional_number(getattr(cut, "cumulative_mole_fraction", None)),
                _optional_number(getattr(cut, "cumulative_mass_fraction", None)),
            ]
            for cut in r.cuts
        ],
        number_formats=[None, _FMT_INTEGER, _FMT_DIMENSIONLESS_HIGH, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW, _FMT_TEMPERATURE, _FMT_DIMENSIONLESS_LOW, _FMT_DIMENSIONLESS_LOW],
    )
